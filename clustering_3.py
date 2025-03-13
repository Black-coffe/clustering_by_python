import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import numpy as np
from collections import defaultdict, Counter
import re
import threading
from ttkthemes import ThemedTk
import mysql.connector
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd
import json
from datetime import datetime
from config import DB_CONFIG
import zipfile
import os
import shutil
from collections import Counter
import tempfile


# ===================== Tooltip Class =====================
class ToolTip:
    """
    Class for creating tooltips on GUI elements
    """

    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)
        self.timer_id = None

    def show_tooltip(self, event=None):
        self.timer_id = self.widget.after(1000, self._show_tooltip)

    def _show_tooltip(self):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25

        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")

        label = tk.Label(self.tooltip, text=self.text, background="#FFFFDD",
                         relief="solid", borderwidth=1, wraplength=350,
                         justify="left", padx=5, pady=5)
        label.pack()

    def hide_tooltip(self, event=None):
        if self.timer_id:
            self.widget.after_cancel(self.timer_id)
            self.timer_id = None
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None


# ===================== Global Variables =====================
root = None
status_bar = None
progress_bar = None
progress_text = None

# Clustering parameters
threshold_value = None
max_position = None
cluster_type = None
excluded_domains_text = None
min_frequency = None
prefix_filter = None
export_detail = None

# UI components
clusters_tree = None
details_text = None
search_var = None
stats_text = None
prefix_combo = None

# Data storage
current_clusters = None
current_data = None
current_keywords_data = None
current_serp_data = None
current_related_data = None
current_paa_data = None
keywords_by_prefix = None
stop_clustering = False

# Constants
DEFAULT_EXCLUDED_DOMAINS = ['wikipedia.org', 'youtube.com', 'facebook.com', 'instagram.com']


# ===================== Data Loading and Processing Functions =====================

def load_all_data_from_db():
    """
    Loads all necessary data from the database:
    - keywords and frequency
    - search queries and results
    - related queries and people also ask
    """
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        # 1. Loading keywords with frequency and prefixes
        query_keywords = """
            SELECT id, query, frequency, prefix 
            FROM keywords
        """
        cursor.execute(query_keywords)
        keywords_data = {row['query']: row for row in cursor.fetchall()}

        # 2. Loading search queries with metadata
        query_queries = """
            SELECT id, query, timestamp, country, language
            FROM queries
        """
        cursor.execute(query_queries)
        queries_data = {row['query']: row for row in cursor.fetchall()}

        # 3. Loading organic results
        query_organic = """
            SELECT q.query, o.position, o.url, o.title, o.description
            FROM queries q 
            JOIN organic_results o ON q.id = o.query_id
            ORDER BY o.position
        """
        cursor.execute(query_organic)

        serp_data = defaultdict(list)
        for row in cursor.fetchall():
            normalized_url = normalize_url(row['url']) if row['url'] else None
            if normalized_url:
                serp_data[row['query']].append({
                    'position': row['position'],
                    'url': row['url'],
                    'normalized_url': normalized_url,
                    'title': row['title'],
                    'description': row['description']
                })

        # 4. Loading related queries
        query_related = """
            SELECT q.query, r.related_query
            FROM queries q 
            JOIN related_queries r ON q.id = r.query_id
        """
        cursor.execute(query_related)
        related_data = defaultdict(list)
        for row in cursor.fetchall():
            related_data[row['query']].append(row['related_query'])

        # 5. Loading "People also ask"
        query_paa = """
            SELECT q.query, p.question
            FROM queries q 
            JOIN people_also_ask p ON q.id = p.query_id
        """
        cursor.execute(query_paa)
        paa_data = defaultdict(list)
        for row in cursor.fetchall():
            paa_data[row['query']].append(row['question'])

        conn.close()
        return keywords_data, serp_data, queries_data, related_data, paa_data

    except mysql.connector.Error as e:
        messagebox.showerror("Database Error", f"Failed to load data: {e}")
        return {}, {}, {}, {}, {}


def normalize_url(url):
    """
    Normalizes URLs for more accurate comparison:
    - Removes protocol (http/https)
    - Optionally removes www and subdomains
    - Optionally cleans query parameters and anchors
    """
    if not url:
        return None

    try:
        parsed = urlparse(url)
        normalized = parsed.netloc + parsed.path

        if normalized.startswith('www.'):
            normalized = normalized[4:]

        if normalized.endswith('/'):
            normalized = normalized[:-1]

        return normalized.lower()
    except:
        return url.lower()


def filter_domains(serp_data, excluded_domains):
    """
    Filters SERP results, excluding specified domains
    """
    filtered_serp = {}

    for query, results in serp_data.items():
        filtered_results = []
        for result in results:
            if result['url']:
                domain = urlparse(result['url']).netloc
                if not any(ex_domain in domain for ex_domain in excluded_domains):
                    filtered_results.append(result)

        if filtered_results:  # Only if results remain after filtering
            filtered_serp[query] = filtered_results

    return filtered_serp


def prepare_url_sets(serp_data, max_position=None):
    """
    Prepares URL sets for each keyword, with optional position limit
    """
    url_sets = {}

    for query, results in serp_data.items():
        if max_position:
            url_set = set(result['normalized_url'] for result in results
                          if result['position'] <= max_position and result['normalized_url'])
        else:
            url_set = set(result['normalized_url'] for result in results if result['normalized_url'])

        if url_set:  # Add only if there's at least one URL
            url_sets[query] = url_set

    return url_sets


def group_keywords_by_prefix(keywords_data):
    """
    Groups keywords by prefixes for preliminary analysis
    """
    grouped = defaultdict(list)

    for query, data in keywords_data.items():
        prefix = data.get('prefix', 'No prefix')
        grouped[prefix].append(query)

    return grouped


# ===================== Core Clustering Functions =====================

def get_central_keyword(cluster, keywords_data):
    """
    Selects the central keyword based on frequency
    """
    if not cluster or not keywords_data:
        return None

    # Find the keyword with highest frequency
    central = None
    max_freq = 0

    for kw in cluster:
        if kw in keywords_data:
            freq = keywords_data[kw].get('frequency', 0)
            if freq > max_freq:
                max_freq = freq
                central = kw

    # If no frequency data, take the first keyword
    if not central and cluster:
        central = next(iter(cluster))

    return central


def serp_clustering(serp_data, threshold, mode="soft", max_position=10):
    """
    SERP-based clustering with configurable parameters

    Parameters:
    - serp_data: Dictionary of SERP results
    - threshold: Minimum number of common URLs
    - mode: "soft" or "hard"
    - max_position: Maximum position to consider in SERP

    Returns:
    - List of clusters (sets of keywords)
    """
    global stop_clustering
    stop_clustering = False

    # Prepare URL sets for comparison
    url_sets = prepare_url_sets(serp_data, max_position)
    clusters = []
    remaining = set(url_sets.keys())

    # Main clustering loop
    while remaining and not stop_clustering:
        # Choose the highest frequency keyword as a seed
        if current_keywords_data:
            seed_keyword = max(remaining,
                               key=lambda kw: current_keywords_data.get(kw, {}).get('frequency', 0)
                               if kw in current_keywords_data else 0)
        else:
            seed_keyword = next(iter(remaining))

        current_cluster = {seed_keyword}

        # For soft clustering: add keywords with sufficient overlap with the seed
        if mode == "soft":
            for kw in list(remaining - {seed_keyword}):
                # Count common URLs
                intersection = len(url_sets[seed_keyword] & url_sets[kw])

                # Add to cluster if intersection >= threshold
                if intersection >= threshold:
                    current_cluster.add(kw)

        # For hard clustering: each keyword must overlap with all others in cluster
        elif mode == "hard":
            for kw in list(remaining - {seed_keyword}):
                # First check overlap with seed
                if len(url_sets[seed_keyword] & url_sets[kw]) >= threshold:
                    # Then check overlap with all other keywords in the cluster
                    if all(len(url_sets[other] & url_sets[kw]) >= threshold
                           for other in current_cluster if other != kw):
                        current_cluster.add(kw)

        # Add the cluster and update remaining keywords
        if len(current_cluster) > 1:  # Only add if cluster has more than one keyword
            clusters.append(current_cluster)
            remaining -= current_cluster
        elif len(current_cluster) == 1 and seed_keyword in remaining:
            clusters.append(current_cluster)
            remaining.remove(seed_keyword)

    # Handle any remaining isolated keywords
    if remaining and not stop_clustering:
        for kw in remaining:
            clusters.append({kw})

    return clusters


def evaluate_clusters(clusters, url_sets):
    """
    Evaluates clustering quality based on intra-cluster similarity
    and inter-cluster difference
    """
    if not clusters:
        return {
            'avg_intra_sim': 0,
            'avg_inter_sim': 0,
            'silhouette': 0,
            'cluster_stats': []
        }

    # Intra-cluster similarity
    intra_similarities = []
    for cluster in clusters:
        if len(cluster) <= 1:
            continue

        # Calculate average Jaccard similarity between all pairs in cluster
        similarities = []
        for kw1 in cluster:
            for kw2 in cluster:
                if kw1 != kw2:
                    if kw1 in url_sets and kw2 in url_sets:
                        intersection = len(url_sets[kw1] & url_sets[kw2])
                        union = len(url_sets[kw1] | url_sets[kw2])
                        similarity = intersection / union if union > 0 else 0
                        similarities.append(similarity)

        if similarities:
            intra_similarities.append(np.mean(similarities))

    avg_intra_sim = np.mean(intra_similarities) if intra_similarities else 0

    # Inter-cluster similarity
    inter_similarities = []
    for i, cluster1 in enumerate(clusters):
        for j, cluster2 in enumerate(clusters):
            if i >= j:
                continue

            # Calculate average Jaccard similarity between pairs from different clusters
            similarities = []
            for kw1 in cluster1:
                for kw2 in cluster2:
                    if kw1 in url_sets and kw2 in url_sets:
                        intersection = len(url_sets[kw1] & url_sets[kw2])
                        union = len(url_sets[kw1] | url_sets[kw2])
                        similarity = intersection / union if union > 0 else 0
                        similarities.append(similarity)

            if similarities:
                inter_similarities.append(np.mean(similarities))

    avg_inter_sim = np.mean(inter_similarities) if inter_similarities else 0

    # Simplified silhouette (difference between intra-cluster similarity and inter-cluster)
    silhouette = avg_intra_sim - avg_inter_sim

    # Cluster statistics
    cluster_stats = []
    for i, cluster in enumerate(clusters):
        if not cluster:
            continue

        # Average frequency of keywords in cluster
        avg_freq = 0
        total_freq = 0
        if current_keywords_data:
            freqs = [current_keywords_data.get(kw, {}).get('frequency', 0) for kw in cluster]
            avg_freq = np.mean(freqs) if freqs else 0
            total_freq = np.sum(freqs) if freqs else 0

        # Intra-cluster similarity
        intra_sim = 0
        if len(cluster) > 1:
            similarities = []
            for kw1 in cluster:
                for kw2 in cluster:
                    if kw1 != kw2 and kw1 in url_sets and kw2 in url_sets:
                        intersection = len(url_sets[kw1] & url_sets[kw2])
                        union = len(url_sets[kw1] | url_sets[kw2])
                        similarity = intersection / union if union > 0 else 0
                        similarities.append(similarity)
            intra_sim = np.mean(similarities) if similarities else 0

        # Central keyword (highest frequency)
        central = get_central_keyword(cluster, current_keywords_data)

        cluster_stats.append({
            'id': i + 1,
            'size': len(cluster),
            'avg_frequency': avg_freq,
            'total_frequency': total_freq,
            'intra_similarity': intra_sim,
            'central_keyword': central
        })

    return {
        'avg_intra_sim': avg_intra_sim,
        'avg_inter_sim': avg_inter_sim,
        'silhouette': silhouette,
        'cluster_stats': cluster_stats
    }


def detect_common_urls(clusters, serp_data):
    """
    Identifies common URLs for each cluster
    """
    common_urls = {}

    for i, cluster in enumerate(clusters):
        if len(cluster) <= 1:
            continue

        # Collect URLs from all keywords in the cluster
        url_counts = Counter()
        for kw in cluster:
            if kw in serp_data:
                for result in serp_data[kw]:
                    if result['normalized_url']:
                        url_counts[result['normalized_url']] += 1

        # Find URLs present in at least half the keywords
        threshold = max(2, len(cluster) / 2)
        cluster_common_urls = [url for url, count in url_counts.items() if count >= threshold]

        if cluster_common_urls:
            common_urls[i + 1] = cluster_common_urls

    return common_urls


def analyze_keyword_semantics(clusters, related_data, paa_data):
    """
    Analyzes semantic coherence of clusters based on related queries
    and "People also ask" blocks
    """
    semantic_analysis = {}

    for i, cluster in enumerate(clusters):
        cluster_id = i + 1
        semantic_analysis[cluster_id] = {
            'related_queries': [],
            'questions': [],
            'common_words': Counter()
        }

        # Collect all related queries for the cluster
        for kw in cluster:
            if kw in related_data:
                semantic_analysis[cluster_id]['related_queries'].extend(related_data[kw])

            if kw in paa_data:
                semantic_analysis[cluster_id]['questions'].extend(paa_data[kw])

            # Analyze word frequency in queries
            words = re.findall(r'\b\w+\b', kw.lower())
            semantic_analysis[cluster_id]['common_words'].update(words)

        # Keep most frequent related queries and questions
        semantic_analysis[cluster_id]['related_queries'] = list(Counter(
            semantic_analysis[cluster_id]['related_queries']).most_common(10))
        semantic_analysis[cluster_id]['questions'] = list(Counter(
            semantic_analysis[cluster_id]['questions']).most_common(10))

        # Remove stop words from common words
        stop_words = {'and', 'in', 'on', 'with', 'for', 'from', 'to', 'at', 'by', 'of', 'the', 'a', 'an', 'is', 'are'}
        for word in stop_words:
            if word in semantic_analysis[cluster_id]['common_words']:
                del semantic_analysis[cluster_id]['common_words'][word]

        semantic_analysis[cluster_id]['common_words'] = dict(
            semantic_analysis[cluster_id]['common_words'].most_common(15))

    return semantic_analysis


# ===================== Export Functions =====================

def export_clusters_to_excel(clusters, serp_data, filename, include_details=True):
    """
    Exports clusters to Excel with enhanced information and formatting designed for SEO specialists
    """
    # Prepare data for export
    rows = []

    # Track total number of keywords for statistics
    total_keywords = sum(len(cluster) for cluster in clusters)

    # Step 1: Prepare main cluster data
    for idx, cluster in enumerate(clusters, start=1):
        # Determine the "central" keyword (highest frequency)
        central_keyword = get_central_keyword(cluster, current_keywords_data)

        # Count common URLs in the cluster
        url_counts = Counter()
        for kw in cluster:
            if kw in serp_data:
                for result in serp_data[kw]:
                    if result['normalized_url']:
                        url_counts[result['normalized_url']] += 1

        # Find most common URLs
        threshold = max(2, len(cluster) / 2)
        common_urls = [url for url, count in url_counts.most_common() if count >= threshold]
        common_urls_str = "\n".join([url for url in common_urls[:10] if url]) if common_urls else ""

        # Analyze semantic similarity of the cluster
        semantic_analysis = None
        if current_related_data or current_paa_data:
            try:
                semantic_analysis = analyze_keyword_semantics([cluster],
                                                              current_related_data,
                                                              current_paa_data)[1]
            except:
                # Handle any errors in semantic analysis
                semantic_analysis = {
                    'common_words': {},
                    'related_queries': [],
                    'questions': []
                }

        # Calculate cluster statistics
        total_freq = sum(current_keywords_data.get(kw, {}).get('frequency', 0)
                         for kw in cluster if kw in current_keywords_data)

        avg_freq = total_freq / len(cluster) if cluster else 0

        # Add rows for each keyword in the cluster
        for i, kw in enumerate(sorted(cluster,
                                      key=lambda k: current_keywords_data.get(k, {}).get('frequency', 0)
                                      if k in current_keywords_data else 0,
                                      reverse=True)):
            # Get keyword data with safety checks
            frequency = current_keywords_data.get(kw, {}).get('frequency', 0) if current_keywords_data else 0
            prefix = current_keywords_data.get(kw, {}).get('prefix', '') if current_keywords_data else ''

            # Get URLs for this keyword
            kw_urls = []
            if kw in serp_data:
                kw_urls = [result['normalized_url'] for result in serp_data[kw] if result['normalized_url']]

            # Determine if this is the first row of the cluster (for header)
            is_first_in_cluster = (i == 0)

            # Create row for export with safe values
            row = {
                "Cluster ID": idx if is_first_in_cluster else "",
                "Total in cluster": len(cluster) if is_first_in_cluster else "",
                "∑ Frequency": total_freq if is_first_in_cluster else "",
                "Average frequency": round(avg_freq, 2) if is_first_in_cluster else "",
                "Keyword": kw or "",
                "Frequency": frequency or 0,
                "Prefix": prefix or "",
                "Rank in cluster": i + 1,
                "Central keyword": "✓" if kw == central_keyword else ""
            }

            # If first row of cluster, add additional information
            if is_first_in_cluster:
                row["Common URLs in cluster"] = common_urls_str or ""

                # Add query words if semantic analysis exists
                if semantic_analysis and semantic_analysis.get('common_words'):
                    row["Common words"] = ", ".join(
                        [word for word in list(semantic_analysis['common_words'].keys())[:10] if word]) or ""
                else:
                    row["Common words"] = ""

                # Add most frequent related queries with safety checks
                if semantic_analysis and semantic_analysis.get('related_queries'):
                    related_queries = [q[0] for q in semantic_analysis['related_queries'][:5] if q and q[0]]
                    row["Related queries (common)"] = "\n".join(related_queries) if related_queries else ""
                else:
                    row["Related queries (common)"] = ""

                # Add most frequent "People also ask" questions with safety checks
                if semantic_analysis and semantic_analysis.get('questions'):
                    paa_questions = [q[0] for q in semantic_analysis['questions'][:5] if q and q[0]]
                    row["People also ask (common)"] = "\n".join(paa_questions) if paa_questions else ""
                else:
                    row["People also ask (common)"] = ""
            else:
                row["Common URLs in cluster"] = ""
                row["Common words"] = ""
                row["Related queries (common)"] = ""
                row["People also ask (common)"] = ""

            # If details needed, add additional information
            if include_details:
                # Add top URLs for this keyword with safety checks
                for j, url in enumerate(kw_urls[:5], 1):
                    row[f"URL {j}"] = url or ""

                # Add related queries if available
                if current_related_data and kw in current_related_data:
                    related = [q for q in current_related_data[kw][:3] if q]  # Limit to 3 queries
                    row["Related queries"] = "\n".join(related) if related else ""
                else:
                    row["Related queries"] = ""

                # Add "People also ask" if available
                if current_paa_data and kw in current_paa_data:
                    questions = [q for q in current_paa_data[kw][:3] if q]  # Limit to 3 questions
                    row["People also ask"] = "\n".join(questions) if questions else ""
                else:
                    row["People also ask"] = ""

            rows.append(row)

    # Step 2: Prepare top URLs data
    all_urls = []
    for query, results in serp_data.items():
        for result in results:
            if result.get('normalized_url'):
                all_urls.append({
                    'url': result['normalized_url'],
                    'position': result.get('position', 0),
                    'keyword': query,
                    'title': result.get('title', ''),
                    'description': result.get('description', '')
                })

    # Count URL frequencies
    url_counter = Counter([item['url'] for item in all_urls])
    top_urls = [{'url': url, 'count': count} for url, count in url_counter.most_common(100)]

    # Add additional data to top URLs
    for url_data in top_urls:
        # Find a sample keyword for this URL
        sample_items = [item for item in all_urls if item['url'] == url_data['url']]
        if sample_items:
            sample = sample_items[0]
            url_data['sample_keyword'] = sample['keyword']
            url_data['sample_title'] = sample['title']
            url_data['sample_description'] = sample.get('description', '')[:100] + "..." if sample.get(
                'description') else ""

            # Calculate average position
            positions = [item['position'] for item in sample_items if item['position']]
            url_data['avg_position'] = sum(positions) / len(positions) if positions else 0

            # Calculate number of keywords
            keywords_with_url = set(item['keyword'] for item in sample_items)
            url_data['keyword_count'] = len(keywords_with_url)
        else:
            url_data.update({
                'sample_keyword': '',
                'sample_title': '',
                'sample_description': '',
                'avg_position': 0,
                'keyword_count': 0
            })

    # Function to remove None values
    def replace_none(data):
        if isinstance(data, dict):
            return {k: replace_none(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [replace_none(item) for item in data]
        elif data is None:
            return ""
        else:
            return data

    # Clean data
    rows = replace_none(rows)
    top_urls = replace_none(top_urls)

    # Create DataFrame and save to Excel with safety guards
    try:
        # Create main DataFrame
        df = pd.DataFrame(rows)

        # Create top URLs DataFrame
        top_urls_df = pd.DataFrame(top_urls)

        # Configure Excel writing
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Write main cluster data
            df.to_excel(writer, index=False, sheet_name="Clusters")

            # Write top URLs data
            if top_urls_df.shape[0] > 0:
                top_urls_df.to_excel(writer, index=False, sheet_name="Top 100 URLs")

            # Add clustering statistics to additional sheet
            if include_details and current_serp_data:
                url_sets = prepare_url_sets(current_serp_data)
                eval_results = evaluate_clusters(clusters, url_sets)

                # Prepare cluster stats
                stats_data = []
                for stat in eval_results['cluster_stats']:
                    stats_data.append({
                        "Cluster ID": stat['id'],
                        "Size": stat['size'],
                        "Average frequency": stat['avg_frequency'] if stat.get('avg_frequency') is not None else 0,
                        "Total frequency": stat['total_frequency'] if stat.get('total_frequency') is not None else 0,
                        "Internal similarity": stat['intra_similarity'] if stat.get(
                            'intra_similarity') is not None else 0,
                        "Central keyword": stat['central_keyword'] if stat.get('central_keyword') is not None else ""
                    })

                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, index=False, sheet_name="Statistics")

                # Add general statistics
                summary_data = {
                    "Metric": ["Number of clusters", "Total keywords", "Average cluster size",
                               "Average intra-cluster similarity", "Average inter-cluster similarity",
                               "Silhouette (clustering quality)"],
                    "Value": [
                        len(clusters),
                        total_keywords,
                        total_keywords / len(clusters) if len(clusters) > 0 else 0,
                        eval_results.get('avg_intra_sim', 0),
                        eval_results.get('avg_inter_sim', 0),
                        eval_results.get('silhouette', 0)
                    ]
                }

                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, index=False, sheet_name="Summary", startrow=0)

        # Step 3: Apply rich formatting using openpyxl
        wb = load_workbook(filename)

        # CLUSTERS SHEET FORMATTING
        if "Clusters" in wb.sheetnames:
            ws = wb["Clusters"]

            # Create styles for headers and grouping
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cluster_header_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
            alternate_fill = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
            highlight_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            central_kw_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Apply styles to headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            # Set optimal column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)  # Maximum 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Format data by rows and set up individual cluster grouping
            prev_cluster = None
            group_start_row = None

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                cluster_id = ws.cell(row=row_idx, column=1).value
                is_central = ws.cell(row=row_idx, column=9).value == "✓"  # Check for central keyword

                # Create new group when cluster ID changes
                if cluster_id and (prev_cluster != cluster_id or prev_cluster is None):
                    # If not the first cluster, end the previous group
                    if prev_cluster is not None and group_start_row is not None:
                        # Create an outline group for each cluster
                        ws.row_dimensions.group(group_start_row, row_idx - 1, outline_level=1)
                        # Set the rows to be hidden initially (collapsed)
                        for hidden_row in range(group_start_row + 1, row_idx):
                            ws.row_dimensions[hidden_row].hidden = True

                    # Start new group with the header row
                    group_start_row = row_idx

                    # Style for new cluster group
                    for cell in row:
                        cell.fill = cluster_header_fill
                        cell.font = Font(bold=True)
                        cell.border = thin_border

                    prev_cluster = cluster_id
                # Style for central keyword
                elif is_central:
                    for cell in row:
                        cell.fill = central_kw_fill
                        cell.border = thin_border
                # Alternate rows for readability
                elif row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = alternate_fill
                        cell.border = thin_border
                else:
                    for cell in row:
                        cell.border = thin_border

                # Alignment and text wrapping
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.alignment = Alignment(horizontal='right', vertical='top')

            # Create the final group for the last cluster
            if prev_cluster is not None and group_start_row is not None and row_idx >= group_start_row:
                ws.row_dimensions.group(group_start_row, row_idx, outline_level=1)
                # Hide rows in the last group too
                for hidden_row in range(group_start_row + 1, row_idx + 1):
                    ws.row_dimensions[hidden_row].hidden = True

            # Add conditional formatting to highlight high frequency
            from openpyxl.formatting.rule import CellIsRule
            freq_col = 6  # Frequency column
            ws.conditional_formatting.add(f"{get_column_letter(freq_col)}2:{get_column_letter(freq_col)}{ws.max_row}",
                                          CellIsRule(operator='greaterThan', formula=['1000'],
                                                     stopIfTrue=True, fill=highlight_fill))

        # TOP URLS SHEET FORMATTING
        if "Top 100 URLs" in wb.sheetnames:
            ws_urls = wb["Top 100 URLs"]

            # Format headers
            for cell in ws_urls[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            # Set column widths
            ws_urls.column_dimensions['A'].width = 50  # URL column
            ws_urls.column_dimensions['B'].width = 10  # Count column
            ws_urls.column_dimensions['C'].width = 30  # Sample keyword
            ws_urls.column_dimensions['D'].width = 40  # Sample title
            ws_urls.column_dimensions['E'].width = 50  # Sample description
            ws_urls.column_dimensions['F'].width = 15  # Average position
            ws_urls.column_dimensions['G'].width = 15  # Keyword count

            # Format data
            for i, row in enumerate(ws_urls.iter_rows(min_row=2), start=2):
                fill = alternate_fill if i % 2 == 0 else None
                for cell in row:
                    if fill:
                        cell.fill = fill
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.alignment = Alignment(horizontal='right', vertical='top')
                    else:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Add position color gradient
            from openpyxl.formatting.rule import ColorScaleRule
            ws_urls.conditional_formatting.add(f"F2:F{ws_urls.max_row}",
                                               ColorScaleRule(start_type='min', start_color='63BE7B',  # Green
                                                              mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                                              # Yellow
                                                              end_type='max', end_color='F8696B'))  # Red

        # STATISTICS SHEET FORMATTING
        if "Statistics" in wb.sheetnames:
            ws_stats = wb["Statistics"]
            # Format headers
            for cell in ws_stats[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Format data
            for i, row in enumerate(ws_stats.iter_rows(min_row=2), start=2):
                fill = alternate_fill if i % 2 == 0 else None
                for cell in row:
                    if fill:
                        cell.fill = fill
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.alignment = Alignment(horizontal='right')

            # Conditional formatting for internal similarity
            ws_stats.conditional_formatting.add(f"E2:E{ws_stats.max_row}",
                                                ColorScaleRule(start_type='min', start_color='F8696B',  # Red
                                                               mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                                               # Yellow
                                                               end_type='max', end_color='63BE7B'))  # Green

        # SUMMARY SHEET FORMATTING
        if "Summary" in wb.sheetnames:
            ws_summary = wb["Summary"]
            # Format headers
            for cell in ws_summary[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Format data
            for row in ws_summary.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.alignment = Alignment(horizontal='right')
                        cell.font = Font(bold=True)
                    else:
                        cell.font = Font(bold=True)

        # Add autofilters for all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.auto_filter.ref = ws.dimensions

        # Freeze headers
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.freeze_panes = 'A2'

        # Collapse groups initially for cleaner view
        if "Clusters" in wb.sheetnames:
            ws = wb["Clusters"]
            ws.sheet_properties.outlinePr.summaryBelow = False

        # Add the new Ukrainian instruction sheet
        # Add this before the final wb.save(filename) line:

        # Create Ukrainian instruction sheet
        instruction_sheet = wb.create_sheet("Інструкція")

        # Define styles for the instruction sheet
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=12)
        normal_font = Font(size=11)
        highlight_font = Font(bold=True, color="4472C4", size=11)

        # Set column widths
        instruction_sheet.column_dimensions['A'].width = 120

        # Title
        instruction_sheet['A1'].font = title_font
        instruction_sheet['A1'] = "Інструкція з використання звіту кластеризації ключових слів для SEO-спеціаліста"

        # Introduction
        row = 3
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "Цей файл Excel містить результати кластеризації пошукових запитів (ключових слів) на основі аналізу пошукової видачі Google (SERP). Кластеризація допомагає згрупувати семантично пов'язані запити для створення оптимальної структури сайту та покращення SEO."
        row += 2

        # Sheet descriptions
        instruction_sheet[f'A{row}'].font = subtitle_font
        instruction_sheet[f'A{row}'] = "ОПИС ВКЛАДОК ФАЙЛУ:"
        row += 2

        # Clusters sheet
        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "Вкладка 'Clusters' (Кластери):"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "Основна вкладка, що містить всі сформовані кластери ключових слів. Кожен кластер представлений як окрема група, яку можна розгорнути або згорнути натисканням на знак '+/-' зліва."
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "Основні колонки:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Cluster ID - унікальний номер кластера"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Total in cluster - загальна кількість ключових слів у кластері"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• ∑ Frequency - сумарна частотність (пошуковий об'єм) всіх запитів у кластері"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Average frequency - середня частотність запитів у кластері"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Keyword - текст ключового слова/пошукового запиту"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Frequency - індивідуальна частотність (пошуковий об'єм) кожного запиту"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Prefix - префікс або категорія запиту (якщо вказано в базі даних)"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Rank in cluster - ранг запиту в кластері (сортування за частотністю)"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Central keyword - позначка '✓' для центрального запиту кластера (найбільш репрезентативний запит з найвищою частотністю)"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Common URLs in cluster - загальні URL, які присутні в результатах пошуку для більшості запитів кластера"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Common words - найчастіші слова, які зустрічаються в запитах кластера"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Related queries (common) - пов'язані пошукові запити, характерні для кластера"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• People also ask (common) - питання з блоку 'Люди також питають', характерні для кластера"
        row += 2

        # Top 100 URLs sheet
        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "Вкладка 'Top 100 URLs' (Топ-100 URL):"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "Містить список із 100 найбільш поширених URL, які зустрічаються в результатах пошуку для проаналізованих ключових слів."
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "Основні колонки:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• url - нормалізований URL-адреса (без протоколу та параметрів)"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• count - кількість запитів, в пошуковій видачі яких зустрічається цей URL"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• sample_keyword - приклад ключового слова, для якого цей URL був у видачі"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• sample_title - приклад заголовка сторінки в результатах пошуку"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• sample_description - приклад опису (сніппета) сторінки"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• avg_position - середня позиція URL в пошуковій видачі (колір від зеленого до червоного: зелений - найкращі позиції, червоний - найгірші)"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• keyword_count - кількість різних ключових слів, для яких цей URL був у пошуковій видачі"
        row += 2

        # Statistics sheet
        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "Вкладка 'Statistics' (Статистика):"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "Містить детальну статистику по кожному кластеру."
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "Основні колонки:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Cluster ID - ідентифікатор кластера"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Size - розмір кластера (кількість ключових слів)"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Average frequency - середня частотність запитів у кластері"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Total frequency - загальна сумарна частотність кластера"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Internal similarity - внутрішня подібність запитів у кластері (від 0 до 1, де 1 - максимальна подібність). Колір від червоного до зеленого: зелений - висока подібність, червоний - низька."
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Central keyword - центральний запит кластера"
        row += 2

        # Summary sheet
        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "Вкладка 'Summary' (Зведення):"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "Містить загальну інформацію про результати кластеризації."
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "Включає такі метрики:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Number of clusters - загальна кількість створених кластерів"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Total keywords - загальна кількість проаналізованих ключових слів"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Average cluster size - середній розмір кластера"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Average intra-cluster similarity - середня внутрішня подібність запитів у кластерах"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Average inter-cluster similarity - середня подібність між кластерами (бажано низьке значення)"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Silhouette (clustering quality) - силует (якість кластеризації): від -1 до 1, де значення ближче до 1 вказує на кращу якість кластеризації"
        row += 2

        # How to use section
        instruction_sheet[f'A{row}'].font = subtitle_font
        instruction_sheet[f'A{row}'] = "ЯК ВИКОРИСТОВУВАТИ РЕЗУЛЬТАТИ КЛАСТЕРИЗАЦІЇ ДЛЯ SEO:"
        row += 2

        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "1. Створення структури сайту:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Кожен кластер може стати основою для окремої сторінки на сайті"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Центральний запит кластера (Central keyword) найкраще використовувати як основний запит для сторінки"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Інші запити кластера можуть бути додатковими для оптимізації тієї ж сторінки"
        row += 2

        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "2. Створення контенту:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Використовуйте Common words для визначення основних тематичних слів, які потрібно включити в контент"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Related queries та People also ask можуть стати основою для підзаголовків і додаткових розділів на сторінці"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Аналізуйте заголовки та описи конкурентів у Common URLs для розуміння, який контент отримує високі позиції"
        row += 2

        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "3. Аналіз конкурентів:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Вивчайте Common URLs для визначення основних конкурентів по кожному кластеру"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• У вкладці Top 100 URLs проаналізуйте, які сайти найчастіше зустрічаються в видачі та на яких позиціях"
        row += 2

        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "4. Оптимізація існуючого контенту:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Якщо URL вашого сайту вже є в Common URLs для кластера, перевірте, чи оптимізована сторінка під всі запити кластера"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Використовуйте Internal similarity для визначення, наскільки тісно пов'язані запити в кластері - вищі значення означають сильніший зв'язок між запитами"
        row += 2

        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "5. Пріоритизація роботи:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Фокусуйтеся спочатку на кластерах з найвищою сумарною частотністю (∑ Frequency)"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Звертайте увагу на великі кластери з високою внутрішньою подібністю - вони часто представляють чіткі тематичні групи"
        row += 2

        instruction_sheet[f'A{row}'].font = highlight_font
        instruction_sheet[f'A{row}'] = "Практичні поради:"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Використовуйте функцію 'Згорнути/Розгорнути' кластери (знак +/- зліва) для зручності роботи з великим об'ємом даних"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Сортуйте дані за різними колонками для різних видів аналізу (наприклад, за частотністю або розміром кластера)"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[
            f'A{row}'] = "• Звертайте увагу на кольорові індикатори - вони допомагають швидко визначити важливі аспекти даних"
        row += 1
        instruction_sheet[f'A{row}'].font = normal_font
        instruction_sheet[f'A{row}'] = "• Для пошуку конкретного запиту використовуйте функцію пошуку Excel (Ctrl+F)"

        # Format instruction sheet
        for r in range(1, row + 1):
            instruction_sheet[f'A{r}'].alignment = Alignment(wrap_text=True)

        # Add autofilter and freeze panes
        instruction_sheet.freeze_panes = 'A2'

        # Adjust column widths
        instruction_sheet.column_dimensions['A'].width = 120

        # Save workbook
        wb.save(filename)
        return filename

    except Exception as e:
        # Log error and return a more informative error message
        print(f"Excel export error: {str(e)}")
        raise Exception(f"Failed to export to Excel: {str(e)}")


def calculate_domain_stats(serp_data):
    """
    Calculates domain statistics to identify common domains in SERP results
    """
    domain_counter = Counter()

    for query, results in serp_data.items():
        for result in results:
            if result['url']:
                domain = urlparse(result['url']).netloc
                domain_counter[domain] += 1

    return domain_counter


def export_clusters_to_csv(clusters, serp_data, filename):
    """
    Exports clusters to multiple CSV files packed in a ZIP archive.
    Format similar to Excel export with multiple sheets.
    """
    import zipfile
    import os
    import tempfile
    import pandas as pd
    from datetime import datetime
    from urllib.parse import urlparse

    # Create temporary directory for storing CSV files
    temp_dir = tempfile.mkdtemp()

    try:
        # Track total number of keywords for statistics
        total_keywords = sum(len(cluster) for cluster in clusters)

        # 1. Create main Clusters file
        rows = []
        for idx, cluster in enumerate(clusters, start=1):
            # Determine the "central" keyword (highest frequency)
            central_keyword = get_central_keyword(cluster, current_keywords_data)

            # Count common URLs in the cluster
            url_counts = Counter()
            for kw in cluster:
                if kw in serp_data:
                    for result in serp_data[kw]:
                        if result['normalized_url']:
                            url_counts[result['normalized_url']] += 1

            # Find most common URLs
            threshold = max(2, len(cluster) / 2)
            common_urls = [url for url, count in url_counts.most_common() if count >= threshold]
            common_urls_str = "\n".join([url for url in common_urls[:10] if url]) if common_urls else ""

            # Analyze semantic similarity of the cluster
            semantic_analysis = None
            if current_related_data or current_paa_data:
                try:
                    semantic_analysis = analyze_keyword_semantics([cluster],
                                                                  current_related_data,
                                                                  current_paa_data)[1]
                except:
                    semantic_analysis = {
                        'common_words': {},
                        'related_queries': [],
                        'questions': []
                    }

            # Calculate cluster statistics
            total_freq = sum(current_keywords_data.get(kw, {}).get('frequency', 0)
                             for kw in cluster if kw in current_keywords_data)

            avg_freq = total_freq / len(cluster) if cluster else 0

            # Add rows for each keyword in the cluster
            for i, kw in enumerate(sorted(cluster,
                                          key=lambda k: current_keywords_data.get(k, {}).get('frequency', 0)
                                          if k in current_keywords_data else 0,
                                          reverse=True)):
                frequency = current_keywords_data.get(kw, {}).get('frequency', 0) if current_keywords_data else 0
                prefix = current_keywords_data.get(kw, {}).get('prefix', '') if current_keywords_data else ''

                # Get URLs for this keyword
                kw_urls = []
                if kw in serp_data:
                    kw_urls = [result['normalized_url'] for result in serp_data[kw] if result['normalized_url']]

                # Determine if this is the first row of the cluster (for header)
                is_first_in_cluster = (i == 0)

                # Create row for export with safe values
                row = {
                    "Cluster ID": idx if is_first_in_cluster else "",
                    "Total in cluster": len(cluster) if is_first_in_cluster else "",
                    "∑ Frequency": total_freq if is_first_in_cluster else "",
                    "Average frequency": round(avg_freq, 2) if is_first_in_cluster else "",
                    "Keyword": kw or "",
                    "Frequency": frequency or 0,
                    "Prefix": prefix or "",
                    "Rank in cluster": i + 1,
                    "Central keyword": "✓" if kw == central_keyword else ""
                }

                # If first row of cluster, add additional information
                if is_first_in_cluster:
                    row["Common URLs in cluster"] = common_urls_str or ""

                    # Add query words if semantic analysis exists
                    if semantic_analysis and semantic_analysis.get('common_words'):
                        row["Common words"] = ", ".join(
                            [word for word in list(semantic_analysis['common_words'].keys())[:10] if word]) or ""
                    else:
                        row["Common words"] = ""

                    # Add most frequent related queries
                    if semantic_analysis and semantic_analysis.get('related_queries'):
                        related_queries = [q[0] for q in semantic_analysis['related_queries'][:5] if q and q[0]]
                        row["Related queries (common)"] = "\n".join(related_queries) if related_queries else ""
                    else:
                        row["Related queries (common)"] = ""

                    # Add most frequent "People also ask" questions
                    if semantic_analysis and semantic_analysis.get('questions'):
                        paa_questions = [q[0] for q in semantic_analysis['questions'][:5] if q and q[0]]
                        row["People also ask (common)"] = "\n".join(paa_questions) if paa_questions else ""
                    else:
                        row["People also ask (common)"] = ""
                else:
                    row["Common URLs in cluster"] = ""
                    row["Common words"] = ""
                    row["Related queries (common)"] = ""
                    row["People also ask (common)"] = ""

                # Add additional information
                # Add top URLs for this keyword
                for j, url in enumerate(kw_urls[:5], 1):
                    row[f"URL {j}"] = url or ""

                # Add related queries if available
                if current_related_data and kw in current_related_data:
                    related = [q for q in current_related_data[kw][:3] if q]  # Limit to 3 queries
                    row["Related queries"] = "\n".join(related) if related else ""
                else:
                    row["Related queries"] = ""

                # Add "People also ask" if available
                if current_paa_data and kw in current_paa_data:
                    questions = [q for q in current_paa_data[kw][:3] if q]  # Limit to 3 questions
                    row["People also ask"] = "\n".join(questions) if questions else ""
                else:
                    row["People also ask"] = ""

                rows.append(row)

        # Create DataFrame and save to CSV
        df = pd.DataFrame(rows)
        clusters_csv_path = os.path.join(temp_dir, "Clusters.csv")
        df.to_csv(clusters_csv_path, index=False, encoding='utf-8')

        # 2. Prepare top URLs data
        all_urls = []
        for query, results in serp_data.items():
            for result in results:
                if result.get('normalized_url'):
                    all_urls.append({
                        'url': result['normalized_url'],
                        'position': result.get('position', 0),
                        'keyword': query,
                        'title': result.get('title', ''),
                        'description': result.get('description', '')
                    })

        # Count URL frequencies
        url_counter = Counter([item['url'] for item in all_urls])
        top_urls = [{'url': url, 'count': count} for url, count in url_counter.most_common(100)]

        # Add additional data to top URLs
        for url_data in top_urls:
            # Find a sample keyword for this URL
            sample_items = [item for item in all_urls if item['url'] == url_data['url']]
            if sample_items:
                sample = sample_items[0]
                url_data['sample_keyword'] = sample['keyword']
                url_data['sample_title'] = sample['title']
                url_data['sample_description'] = sample.get('description', '')[:100] + "..." if sample.get(
                    'description') else ""

                # Calculate average position
                positions = [item['position'] for item in sample_items if item['position']]
                url_data['avg_position'] = sum(positions) / len(positions) if positions else 0

                # Calculate number of keywords
                keywords_with_url = set(item['keyword'] for item in sample_items)
                url_data['keyword_count'] = len(keywords_with_url)
            else:
                url_data.update({
                    'sample_keyword': '',
                    'sample_title': '',
                    'sample_description': '',
                    'avg_position': 0,
                    'keyword_count': 0
                })

        # Create top URLs DataFrame and save to CSV
        top_urls_df = pd.DataFrame(top_urls)
        top_urls_csv_path = os.path.join(temp_dir, "Top 100 URLs.csv")
        top_urls_df.to_csv(top_urls_csv_path, index=False, encoding='utf-8')

        # 3. Add clustering statistics
        if current_serp_data:
            url_sets = prepare_url_sets(current_serp_data)
            eval_results = evaluate_clusters(clusters, url_sets)

            # Prepare cluster stats
            stats_data = []
            for stat in eval_results['cluster_stats']:
                stats_data.append({
                    "Cluster ID": stat['id'],
                    "Size": stat['size'],
                    "Average frequency": stat['avg_frequency'] if stat.get('avg_frequency') is not None else 0,
                    "Total frequency": stat['total_frequency'] if stat.get('total_frequency') is not None else 0,
                    "Internal similarity": stat['intra_similarity'] if stat.get(
                        'intra_similarity') is not None else 0,
                    "Central keyword": stat['central_keyword'] if stat.get('central_keyword') is not None else ""
                })

            stats_df = pd.DataFrame(stats_data)
            stats_csv_path = os.path.join(temp_dir, "Statistics.csv")
            stats_df.to_csv(stats_csv_path, index=False, encoding='utf-8')

            # Add general statistics
            summary_data = {
                "Metric": ["Number of clusters", "Total keywords", "Average cluster size",
                           "Average intra-cluster similarity", "Average inter-cluster similarity",
                           "Silhouette (clustering quality)"],
                "Value": [
                    len(clusters),
                    total_keywords,
                    total_keywords / len(clusters) if len(clusters) > 0 else 0,
                    eval_results.get('avg_intra_sim', 0),
                    eval_results.get('avg_inter_sim', 0),
                    eval_results.get('silhouette', 0)
                ]
            }

            summary_df = pd.DataFrame(summary_data)
            summary_csv_path = os.path.join(temp_dir, "Summary.csv")
            summary_df.to_csv(summary_csv_path, index=False, encoding='utf-8')

        # 4. Create Ukrainian instruction
        instructions_content = """
Інструкція з використання звіту кластеризації ключових слів для SEO-спеціаліста

Цей файл Excel містить результати кластеризації пошукових запитів (ключових слів) на основі аналізу пошукової видачі Google (SERP). 
Кластеризація допомагає згрупувати семантично пов'язані запити для створення оптимальної структури сайту та покращення SEO.

ОПИС ВКЛАДОК ФАЙЛУ:

Вкладка 'Clusters' (Кластери):
Основна вкладка, що містить всі сформовані кластери ключових слів. Кожен кластер представлений як окрема група.
Основні колонки:
• Cluster ID - унікальний номер кластера
• Total in cluster - загальна кількість ключових слів у кластері
• ∑ Frequency - сумарна частотність (пошуковий об'єм) всіх запитів у кластері
• Average frequency - середня частотність запитів у кластері
• Keyword - текст ключового слова/пошукового запиту
• Frequency - індивідуальна частотність (пошуковий об'єм) кожного запиту
• Prefix - префікс або категорія запиту (якщо вказано в базі даних)
• Rank in cluster - ранг запиту в кластері (сортування за частотністю)
• Central keyword - позначка '✓' для центрального запиту кластера (найбільш репрезентативний запит з найвищою частотністю)
• Common URLs in cluster - загальні URL, які присутні в результатах пошуку для більшості запитів кластера
• Common words - найчастіші слова, які зустрічаються в запитах кластера
• Related queries (common) - пов'язані пошукові запити, характерні для кластера
• People also ask (common) - питання з блоку 'Люди також питають', характерні для кластера

Вкладка 'Top 100 URLs' (Топ-100 URL):
Містить список із 100 найбільш поширених URL, які зустрічаються в результатах пошуку для проаналізованих ключових слів.
Основні колонки:
• url - нормалізований URL-адреса (без протоколу та параметрів)
• count - кількість запитів, в пошуковій видачі яких зустрічається цей URL
• sample_keyword - приклад ключового слова, для якого цей URL був у видачі
• sample_title - приклад заголовка сторінки в результатах пошуку
• sample_description - приклад опису (сніппета) сторінки
• avg_position - середня позиція URL в пошуковій видачі
• keyword_count - кількість різних ключових слів, для яких цей URL був у пошуковій видачі

Вкладка 'Statistics' (Статистика):
Містить детальну статистику по кожному кластеру.
Основні колонки:
• Cluster ID - ідентифікатор кластера
• Size - розмір кластера (кількість ключових слів)
• Average frequency - середня частотність запитів у кластері
• Total frequency - загальна сумарна частотність кластера
• Internal similarity - внутрішня подібність запитів у кластері (від 0 до 1, де 1 - максимальна подібність)
• Central keyword - центральний запит кластера

Вкладка 'Summary' (Зведення):
Містить загальну інформацію про результати кластеризації.
Включає такі метрики:
• Number of clusters - загальна кількість створених кластерів
• Total keywords - загальна кількість проаналізованих ключових слів
• Average cluster size - середній розмір кластера
• Average intra-cluster similarity - середня внутрішня подібність запитів у кластерах
• Average inter-cluster similarity - середня подібність між кластерами (бажано низьке значення)
• Silhouette (clustering quality) - силует (якість кластеризації): від -1 до 1, де значення ближче до 1 вказує на кращу якість кластеризації

ЯК ВИКОРИСТОВУВАТИ РЕЗУЛЬТАТИ КЛАСТЕРИЗАЦІЇ ДЛЯ SEO:

1. Створення структури сайту:
• Кожен кластер може стати основою для окремої сторінки на сайті
• Центральний запит кластера (Central keyword) найкраще використовувати як основний запит для сторінки
• Інші запити кластера можуть бути додатковими для оптимізації тієї ж сторінки

2. Створення контенту:
• Використовуйте Common words для визначення основних тематичних слів, які потрібно включити в контент
• Related queries та People also ask можуть стати основою для підзаголовків і додаткових розділів на сторінці
• Аналізуйте заголовки та описи конкурентів у Common URLs для розуміння, який контент отримує високі позиції

3. Аналіз конкурентів:
• Вивчайте Common URLs для визначення основних конкурентів по кожному кластеру
• У вкладці Top 100 URLs проаналізуйте, які сайти найчастіше зустрічаються в видачі та на яких позиціях

4. Оптимізація існуючого контенту:
• Якщо URL вашого сайту вже є в Common URLs для кластера, перевірте, чи оптимізована сторінка під всі запити кластера
• Використовуйте Internal similarity для визначення, наскільки тісно пов'язані запити в кластері - вищі значення означають сильніший зв'язок між запитами

5. Пріоритизація роботи:
• Фокусуйтеся спочатку на кластерах з найвищою сумарною частотністю (∑ Frequency)
• Звертайте увагу на великі кластери з високою внутрішньою подібністю - вони часто представляють чіткі тематичні групи
"""
        instructions_path = os.path.join(temp_dir, "Інструкція.txt")
        with open(instructions_path, 'w', encoding='utf-8') as f:
            f.write(instructions_content)

        # 5. Create ZIP archive
        zip_path = filename
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    zipf.write(file_path, os.path.basename(file_path))

        return zip_path

    finally:
        # Clean up temporary directory
        import shutil
        shutil.rmtree(temp_dir)


def export_clusters_to_json(clusters, serp_data, filename):
    """
    Exports clusters to JSON format with comprehensive information
    """
    export_data = {
        "clusters": [],
        "metadata": {
            "timestamp": datetime.now().isoformat(),
            "total_clusters": len(clusters),
            "total_keywords": sum(len(c) for c in clusters)
        }
    }

    for idx, cluster in enumerate(clusters, start=1):
        cluster_data = {
            "id": idx,
            "size": len(cluster),
            "keywords": [],
            "common_urls": []
        }

        # Count common URLs in the cluster
        url_counts = Counter()
        for kw in cluster:
            if kw in serp_data:
                for result in serp_data[kw]:
                    if result['normalized_url']:
                        url_counts[result['normalized_url']] += 1

        # Find most common URLs (present in at least half the keywords)
        threshold = max(2, len(cluster) / 2)
        common_urls = [{"url": url, "count": count}
                       for url, count in url_counts.most_common() if count >= threshold]
        cluster_data["common_urls"] = common_urls[:10]  # Limit to 10 URLs

        # Add information about each keyword
        for kw in sorted(cluster):
            keyword_data = {
                "query": kw,
                "frequency": current_keywords_data.get(kw, {}).get('frequency', 0) if current_keywords_data else 0,
                "prefix": current_keywords_data.get(kw, {}).get('prefix', '') if current_keywords_data else '',
                "urls": []
            }

            # Add URLs for this keyword
            if kw in serp_data:
                for result in serp_data[kw]:
                    if result['normalized_url']:
                        keyword_data["urls"].append({
                            "url": result['normalized_url'],
                            "position": result['position'],
                            "title": result['title'],
                            "description": result['description']
                        })

            # Add related queries, if available
            if current_related_data and kw in current_related_data:
                keyword_data["related_queries"] = current_related_data[kw][:5]  # Limit to 5 queries

            # Add "People also ask", if available
            if current_paa_data and kw in current_paa_data:
                keyword_data["people_also_ask"] = current_paa_data[kw][:5]  # Limit to 5 questions

            cluster_data["keywords"].append(keyword_data)

        export_data["clusters"].append(cluster_data)

    # Save to JSON file
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)

    return filename


# ===================== GUI =====================

def initialize_gui():
    """
    Initializes the GUI with simplified controls for SEO clustering
    """
    global root
    global status_bar, progress_bar, progress_text
    global prefix_combo
    global cluster_type, threshold_value
    global excluded_domains_text, max_position, prefix_filter, min_frequency
    global export_detail, clusters_tree, details_text, search_var, details_frame
    global stats_text

    # Create main window with dark theme
    root = ThemedTk(theme="arc")
    root.title("SEO Keyword Clustering")
    root.geometry("1280x800")

    # ========= Top Menu =========
    menu_bar = tk.Menu(root)

    file_menu = tk.Menu(menu_bar, tearoff=0)
    file_menu.add_command(label="Load Data", command=load_data_from_database)
    file_menu.add_separator()
    file_menu.add_command(label="Export to Excel", command=lambda: export_current_clusters("excel"))
    file_menu.add_command(label="Export to CSV", command=lambda: export_current_clusters("csv"))
    file_menu.add_command(label="Export to JSON", command=lambda: export_current_clusters("json"))
    file_menu.add_separator()
    file_menu.add_command(label="Exit", command=root.quit)
    menu_bar.add_cascade(label="File", menu=file_menu)

    clustering_menu = tk.Menu(menu_bar, tearoff=0)
    clustering_menu.add_command(label="Run Clustering", command=run_clustering)
    clustering_menu.add_command(label="Stop Processing", command=stop_processing)
    menu_bar.add_cascade(label="Clustering", menu=clustering_menu)

    help_menu = tk.Menu(menu_bar, tearoff=0)
    help_menu.add_command(label="About", command=show_about)
    help_menu.add_command(label="Help", command=show_help)
    menu_bar.add_cascade(label="Help", menu=help_menu)

    root.config(menu=menu_bar)

    # ========= Main Panel =========
    main_paned = ttk.PanedWindow(root, orient=tk.HORIZONTAL)
    main_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    left_frame = ttk.Frame(main_paned, width=300)
    main_paned.add(left_frame, weight=1)

    right_frame = ttk.Frame(main_paned)
    main_paned.add(right_frame, weight=3)

    # ======= Left Tabs (Settings) =======
    settings_notebook = ttk.Notebook(left_frame)
    settings_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    general_settings_frame = ttk.Frame(settings_notebook)
    settings_notebook.add(general_settings_frame, text="Clustering")

    filtering_frame = ttk.Frame(settings_notebook)
    settings_notebook.add(filtering_frame, text="Filtering")

    # ======= Right Tabs (Results) =======
    results_notebook = ttk.Notebook(right_frame)
    results_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    results_frame = ttk.Frame(results_notebook)
    results_notebook.add(results_frame, text="Clustering Results")

    stats_frame = ttk.Frame(results_notebook)
    results_notebook.add(stats_frame, text="Statistics")

    # ==================== PARAMETERS (Clustering) ====================
    settings_frame = general_settings_frame

    # -- Clustering Type
    ttk.Label(settings_frame, text="Clustering Type:").pack(anchor=tk.W, padx=5, pady=2)
    cluster_type = tk.StringVar(value="soft")
    ttk.Radiobutton(settings_frame, text="Soft", variable=cluster_type, value="soft").pack(anchor=tk.W, padx=15, pady=2)
    ttk.Radiobutton(settings_frame, text="Hard", variable=cluster_type, value="hard").pack(anchor=tk.W, padx=15, pady=2)

    ToolTip(settings_frame,
            "Clustering Type:\n"
            "• Soft: Keywords with shared URLs in SERP are grouped\n"
            "• Hard: Every keyword must share URLs with all others in the cluster")

    # -- URL Overlap Threshold
    ttk.Label(settings_frame, text="URL Overlap Threshold:").pack(anchor=tk.W, padx=5, pady=2)
    threshold_frame = ttk.Frame(settings_frame)
    threshold_frame.pack(fill=tk.X, padx=5, pady=2)

    threshold_value = tk.IntVar(value=3)
    threshold_slider = ttk.Scale(threshold_frame, from_=1, to=10,
                                 variable=threshold_value, orient=tk.HORIZONTAL,
                                 command=lambda v: threshold_spinbox.set(int(float(v))))
    threshold_slider.pack(side=tk.LEFT, fill=tk.X, expand=True)

    threshold_spinbox = ttk.Spinbox(threshold_frame, from_=1, to=10, textvariable=threshold_value, width=3)
    threshold_spinbox.pack(side=tk.RIGHT, padx=5)

    ToolTip(threshold_frame,
            "URL Overlap Threshold — minimum number of URLs that must match\n"
            "for keywords to be grouped together.\n\n"
            "Lower (1-3) → larger clusters\n"
            "Higher (7-10) → smaller clusters with high similarity")

    # -- Maximum SERP Position
    ttk.Label(settings_frame, text="Maximum SERP Position to Analyze:").pack(anchor=tk.W, padx=5, pady=2)
    max_position = tk.IntVar(value=10)
    ttk.Radiobutton(settings_frame, text="Top 10", variable=max_position, value=10).pack(anchor=tk.W, padx=15, pady=2)
    ttk.Radiobutton(settings_frame, text="Top 20", variable=max_position, value=20).pack(anchor=tk.W, padx=15, pady=2)
    ttk.Radiobutton(settings_frame, text="All Positions", variable=max_position, value=100).pack(anchor=tk.W, padx=15,
                                                                                                 pady=2)

    ToolTip(settings_frame,
            "Which positions in SERP to analyze:\n"
            "• Top 10: Analyze only first page results (most relevant)\n"
            "• Top 20: Include second page results\n"
            "• All: Include all positions (may add noise)")

    # -- Run button
    ttk.Button(settings_frame, text="Run Clustering", command=run_clustering, style="Accent.TButton") \
        .pack(fill=tk.X, padx=5, pady=10)

    # -- Stop button
    ttk.Button(settings_frame, text="Stop Processing", command=stop_processing) \
        .pack(fill=tk.X, padx=5, pady=5)

    # -- Export settings
    ttk.Label(settings_frame, text="Export Settings:").pack(anchor=tk.W, padx=5, pady=5)
    export_detail = tk.BooleanVar(value=True)
    ttk.Checkbutton(settings_frame, text="Include detailed information in exports", variable=export_detail) \
        .pack(anchor=tk.W, padx=5, pady=2)

    # ==================== PARAMETERS (Filtering) ====================
    ttk.Label(filtering_frame, text="Excluded Domains:").pack(anchor=tk.W, padx=5, pady=2)
    excluded_domains_text = tk.Text(filtering_frame, height=5, width=30)
    excluded_domains_text.pack(fill=tk.X, padx=5, pady=2)

    # Default domains to exclude
    excluded_domains_text.insert(tk.END, "wikipedia.org\nyoutube.com\nfacebook.com\ninstagram.com")

    ttk.Label(filtering_frame, text="Prefix Filter:").pack(anchor=tk.W, padx=5, pady=2)
    prefix_filter = tk.StringVar(value="All Prefixes")
    prefix_combo = ttk.Combobox(filtering_frame, textvariable=prefix_filter, state='readonly')
    prefix_combo['values'] = ['All Prefixes']
    prefix_combo.pack(fill=tk.X, padx=5, pady=2)

    ttk.Label(filtering_frame, text="Minimum Frequency:").pack(anchor=tk.W, padx=5, pady=2)
    min_frequency = tk.IntVar(value=0)
    min_freq_frame = ttk.Frame(filtering_frame)
    min_freq_frame.pack(fill=tk.X, padx=5, pady=2)

    min_freq_slider = ttk.Scale(min_freq_frame, from_=0, to=1000,
                                variable=min_frequency, orient=tk.HORIZONTAL,
                                command=lambda v: min_freq_value_label.config(text=f"{int(float(v))}"))
    min_freq_slider.pack(side=tk.LEFT, fill=tk.X, expand=True)

    min_freq_value_label = ttk.Label(min_freq_frame, text="0")
    min_freq_value_label.pack(side=tk.RIGHT, padx=5)

    ToolTip(min_freq_frame,
            "Minimum frequency — lower threshold of search volume.\n"
            "0 → all keywords\n100 → remove low-frequency\n500 → only high-frequency")

    # -- Apply Filters button
    ttk.Button(filtering_frame, text="Apply Filters", command=apply_filters) \
        .pack(fill=tk.X, padx=5, pady=10)

    # ==================== CLUSTERING RESULTS ====================
    search_frame = ttk.Frame(results_frame)
    search_frame.pack(fill=tk.X, padx=5, pady=5)

    ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT, padx=5)
    search_var = tk.StringVar()
    search_entry = ttk.Entry(search_frame, textvariable=search_var)
    search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
    search_entry.bind("<Return>", lambda e: search_in_clusters())

    ttk.Button(search_frame, text="Find", command=search_in_clusters).pack(side=tk.LEFT, padx=5)
    ttk.Button(search_frame, text="Reset", command=reset_search).pack(side=tk.LEFT, padx=5)

    tree_frame = ttk.Frame(results_frame)
    tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    clusters_tree = ttk.Treeview(tree_frame, columns=("id", "size", "central", "freq"), show="tree headings")
    clusters_tree.heading("id", text="ID")
    clusters_tree.heading("size", text="Size")
    clusters_tree.heading("central", text="Central Keyword")
    clusters_tree.heading("freq", text="∑ Frequency")

    clusters_tree.column("id", width=50, anchor=tk.CENTER)
    clusters_tree.column("size", width=80, anchor=tk.CENTER)
    clusters_tree.column("central", width=300)
    clusters_tree.column("freq", width=120, anchor=tk.E)

    clusters_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=clusters_tree.yview)
    tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    clusters_tree.configure(yscrollcommand=tree_scroll.set)

    clusters_tree.bind("<<TreeviewSelect>>", on_cluster_select)

    # Используем одну и только одну фрейм для деталей кластера
    details_frame = ttk.LabelFrame(results_frame, text="Cluster Details")
    details_frame.pack(fill=tk.BOTH, expand=False, padx=5, pady=5)

    # Создаем панель для текста и скроллбара
    text_panel = ttk.Frame(details_frame)
    text_panel.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    details_text = tk.Text(text_panel, height=10, wrap=tk.WORD)
    details_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    details_scroll = ttk.Scrollbar(text_panel, orient="vertical", command=details_text.yview)
    details_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    details_text.configure(yscrollcommand=details_scroll.set)

    # ==================== STATISTICS ====================
    stats_controls_frame = ttk.Frame(stats_frame)
    stats_controls_frame.pack(fill=tk.X, padx=5, pady=5)

    ttk.Button(stats_controls_frame, text="Update Statistics", command=update_statistics) \
        .pack(side=tk.LEFT, padx=5)

    stats_display_frame = ttk.Frame(stats_frame)
    stats_display_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    stats_text = tk.Text(stats_display_frame, wrap=tk.WORD)
    stats_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    stats_scroll = ttk.Scrollbar(stats_display_frame, orient="vertical", command=stats_text.yview)
    stats_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    stats_text.configure(yscrollcommand=stats_scroll.set)

    # ==================== Status Bar ====================
    status_frame = ttk.Frame(root)
    status_frame.pack(fill=tk.X, padx=5, pady=2)

    status_bar = ttk.Label(status_frame, text="Ready", anchor=tk.W)
    status_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)

    progress_bar = ttk.Progressbar(status_frame, orient=tk.HORIZONTAL, length=200, mode='determinate')
    progress_bar.pack(side=tk.RIGHT, padx=5)

    progress_text = ttk.Label(status_frame, text="0%")
    progress_text.pack(side=tk.RIGHT, padx=5)

    def enable_text_copying():
        # Enable copy/paste in all Text widgets
        def bind_copy_paste(widget):
            widget.bind("<Control-c>", lambda event: widget.event_generate("<<Copy>>"))
            widget.bind("<Control-v>", lambda event: widget.event_generate("<<Paste>>"))
            widget.bind("<Control-a>", lambda event: widget.event_generate("<<SelectAll>>"))

        # Apply to all text widgets
        bind_copy_paste(details_text)
        bind_copy_paste(stats_text)
        if excluded_domains_text:
            bind_copy_paste(excluded_domains_text)

    # Call this function at the end of initialize_gui
    enable_text_copying()


def load_data_from_database():
    """
    Loads data from the database and prepares it for clustering
    """
    global current_keywords_data, current_serp_data, current_queries_data, current_related_data, current_paa_data
    global keywords_by_prefix, prefix_combo

    status_bar.config(text="Loading data from database...")
    progress_bar['value'] = 0
    progress_text.config(text="0%")

    def process_loading():
        global current_keywords_data, current_serp_data, current_queries_data, current_related_data, current_paa_data
        global keywords_by_prefix

        keywords_data, serp_data, queries_data, related_data, paa_data = load_all_data_from_db()

        # Update global variables
        current_keywords_data = keywords_data
        current_serp_data = serp_data
        current_queries_data = queries_data
        current_related_data = related_data
        current_paa_data = paa_data

        # Group keywords by prefixes
        keywords_by_prefix = group_keywords_by_prefix(keywords_data)

        # Return statistics
        return {
            'keywords': len(keywords_data),
            'serp_entries': sum(len(results) for results in serp_data.values()),
            'queries': len(queries_data),
            'prefixes': len(keywords_by_prefix)
        }

    def on_loading_complete(stats):
        # Update interface after loading
        progress_bar['value'] = 100
        progress_text.config(text="100%")

        # Update prefix dropdown
        prefix_values = ['All Prefixes'] + sorted(keywords_by_prefix.keys())
        prefix_combo['values'] = prefix_values

        # Display loading statistics
        status_message = (f"Loaded: {stats['keywords']} keywords, "
                          f"{stats['serp_entries']} SERP results, "
                          f"{stats['prefixes']} prefixes")
        status_bar.config(text=status_message)

        messagebox.showinfo("Data Loading",
                            f"Data successfully loaded!\n\n"
                            f"Keywords: {stats['keywords']}\n"
                            f"SERP results: {stats['serp_entries']}\n"
                            f"Prefixes: {stats['prefixes']}")

    # Run loading in separate thread
    run_processing_in_thread(process_loading, callback=on_loading_complete, timeout=300)


def apply_filters():
    """
    Applies selected filters to data before clustering
    """
    global current_serp_data

    if not current_serp_data:
        messagebox.showwarning("Warning", "Please load data first!")
        return

    # Get filter parameters
    excluded_domains_list = [domain.strip() for domain in excluded_domains_text.get("1.0", tk.END).split("\n") if
                             domain.strip()]
    max_pos = max_position.get()
    selected_prefix = prefix_filter.get()
    min_freq = min_frequency.get()

    status_bar.config(text="Applying filters...")
    progress_bar['value'] = 0
    progress_text.config(text="0%")

    def process_filtering():
        global current_serp_data

        # Copy original data
        filtered_serp = dict(current_serp_data)

        # 1. Filter by domains
        if excluded_domains_list:
            filtered_serp = filter_domains(filtered_serp, excluded_domains_list)

        # 2. Filter by maximum position
        if max_pos < 100:  # 100 is specified as "all positions"
            for query, results in filtered_serp.items():
                filtered_serp[query] = [r for r in results if r['position'] <= max_pos]

        # 3. Filter by prefix
        if selected_prefix != "All Prefixes" and keywords_by_prefix:
            allowed_keywords = set(keywords_by_prefix.get(selected_prefix, []))
            filtered_serp = {k: v for k, v in filtered_serp.items() if k in allowed_keywords}

        # 4. Filter by minimum frequency
        if min_freq > 0 and current_keywords_data:
            filtered_serp = {k: v for k, v in filtered_serp.items()
                             if k in current_keywords_data and
                             current_keywords_data[k].get('frequency', 0) >= min_freq}

        return filtered_serp

    def on_filtering_complete(filtered_serp):
        global current_serp_data

        # Update data
        current_serp_data = filtered_serp

        # Update interface
        progress_bar['value'] = 100
        progress_text.config(text="100%")

        status_message = f"Filters applied. {len(filtered_serp)} keywords remain."
        status_bar.config(text=status_message)

        messagebox.showinfo("Filtering",
                            f"Filters successfully applied!\n\n"
                            f"Remaining keywords: {len(filtered_serp)}")

    # Run filtering in separate thread
    run_processing_in_thread(process_filtering, callback=on_filtering_complete)


def run_clustering():
    """
    Starts the clustering process with selected parameters
    """
    global current_clusters, current_data

    if not current_serp_data:
        messagebox.showwarning("Warning", "Please load data first!")
        return

    # Get clustering parameters
    mode = cluster_type.get()
    threshold_val = threshold_value.get()
    max_pos = max_position.get()

    # Update interface
    status_bar.config(text="Performing clustering...")
    progress_bar['value'] = 10
    progress_text.config(text="10%")

    # Clear clusters tree
    clusters_tree.delete(*clusters_tree.get_children())
    details_text.delete("1.0", tk.END)

    def process_clustering():
        global current_clusters, current_data

        # 1. Prepare SERP data
        progress_update(20, "Preparing data...")
        url_sets = prepare_url_sets(current_serp_data, max_pos)

        # 2. Run SERP-based clustering
        progress_update(50, "Performing clustering...")
        clusters = serp_clustering(current_serp_data, threshold_val, mode, max_pos)

        # 3. Evaluate clusters
        progress_update(80, "Evaluating clusters...")
        current_clusters = clusters
        current_data = url_sets

        # 4. Prepare data for display
        progress_update(90, "Preparing results...")
        cluster_stats = []

        for i, cluster in enumerate(clusters):
            total_freq = 0
            central_kw = get_central_keyword(cluster, current_keywords_data)

            # Calculate total frequency
            for kw in cluster:
                if kw in current_keywords_data:
                    total_freq += current_keywords_data[kw].get('frequency', 0)

            cluster_stats.append({
                'id': i + 1,
                'size': len(cluster),
                'total_freq': total_freq,
                'central': central_kw,
                'keywords': cluster
            })

        return cluster_stats

    def on_clustering_complete(cluster_stats):
        # Update interface with results
        progress_bar['value'] = 100
        progress_text.config(text="100%")

        # Add clusters to tree
        for stat in cluster_stats:
            cluster_id = f"cluster_{stat['id']}"

            # Add cluster
            clusters_tree.insert("", "end", cluster_id, text=f"Cluster {stat['id']}",
                                 values=(stat['id'], stat['size'], stat['central'], f"{stat['total_freq']:,}"))

            # Add keywords to cluster
            for kw in sorted(stat['keywords']):
                freq = current_keywords_data.get(kw, {}).get('frequency', 0) if current_keywords_data else 0
                clusters_tree.insert(cluster_id, "end", text=kw, values=("", "", kw, f"{freq:,}"))

        # Update statistics
        update_statistics()

        # Update status bar
        status_bar.config(text=f"Clustering complete. Created {len(cluster_stats)} clusters.")

    # Run clustering in separate thread
    run_processing_in_thread(process_clustering, callback=on_clustering_complete)


def progress_update(value, message=None):
    """
    Updates progress indicator and message in status bar
    """
    progress_bar['value'] = value
    progress_text.config(text=f"{value}%")

    if message:
        status_bar.config(text=message)

    # Update UI
    root.update_idletasks()


def stop_processing():
    """
    Stops current processing by setting stop flag
    """
    global stop_clustering
    stop_clustering = True
    messagebox.showinfo("Stop", "Clustering process will be stopped.")


def on_cluster_select(event):
    """
    Handles cluster selection in tree
    """
    selected_items = clusters_tree.selection()
    if not selected_items:
        return

    selected_item = selected_items[0]

    # Check if this is a cluster or keyword
    parent_id = clusters_tree.parent(selected_item)

    if not parent_id:  # This is a cluster
        cluster_id = int(clusters_tree.item(selected_item)['values'][0])
        show_cluster_details(cluster_id - 1)  # -1 because indexing starts at 0
    else:  # This is a keyword
        keyword = clusters_tree.item(selected_item)['text']
        show_keyword_details(keyword)


def show_cluster_details(cluster_index):
    """
    Shows details of selected cluster
    """
    if not current_clusters or cluster_index >= len(current_clusters):
        return

    cluster = current_clusters[cluster_index]

    # Clear details text field
    details_text.delete("1.0", tk.END)

    # Determine central keyword
    central_kw = get_central_keyword(cluster, current_keywords_data)

    # Collect common URLs
    common_urls = None
    if current_serp_data:
        url_counts = Counter()
        for kw in cluster:
            if kw in current_serp_data:
                for result in current_serp_data[kw]:
                    if result['normalized_url']:
                        url_counts[result['normalized_url']] += 1

        # Find URLs present in at least half the keywords
        threshold = max(2, len(cluster) / 2)
        common_urls = [url for url, count in url_counts.most_common(10) if count >= threshold]

    # Analyze semantic similarity
    semantic_analysis = None
    if current_related_data or current_paa_data:
        semantic_analysis = analyze_keyword_semantics([cluster],
                                                      current_related_data,
                                                      current_paa_data)[1]

    # Collect cluster statistics
    total_freq = 0
    if current_keywords_data:
        for kw in cluster:
            if kw in current_keywords_data:
                total_freq += current_keywords_data[kw].get('frequency', 0)

    # Display information
    details_text.insert(tk.END, f"=== CLUSTER {cluster_index + 1} ===\n\n", "header")
    details_text.insert(tk.END, f"Size: {len(cluster)} keywords\n")
    details_text.insert(tk.END, f"Total frequency: {total_freq:,}\n")

    if central_kw:
        central_freq = current_keywords_data.get(central_kw, {}).get('frequency', 0) if current_keywords_data else 0
        details_text.insert(tk.END, f"Central keyword: {central_kw} (frequency: {central_freq:,})\n\n")

    # Display keywords
    details_text.insert(tk.END, "Keywords in cluster:\n", "subheader")
    sorted_keywords = sorted(cluster,
                             key=lambda kw: current_keywords_data.get(kw, {}).get('frequency', 0)
                             if current_keywords_data else 0,
                             reverse=True)

    for kw in sorted_keywords[:20]:  # Limit for readability
        freq = current_keywords_data.get(kw, {}).get('frequency', 0) if current_keywords_data else 0
        details_text.insert(tk.END, f"  {kw} ({freq:,})\n")

    if len(sorted_keywords) > 20:
        details_text.insert(tk.END, f"  ... and {len(sorted_keywords) - 20} more keywords\n")

    details_text.insert(tk.END, "\n")

    # Display common URLs
    if common_urls:
        details_text.insert(tk.END, "Common URLs (present in at least half the keywords):\n", "subheader")
        for i, url in enumerate(common_urls, 1):
            details_text.insert(tk.END, f"  {i}. {url}\n")
        details_text.insert(tk.END, "\n")

    # Display semantic analysis
    if semantic_analysis:
        if semantic_analysis['common_words']:
            details_text.insert(tk.END, "Most frequent words in queries:\n", "subheader")
            for word, count in list(semantic_analysis['common_words'].items())[:10]:
                details_text.insert(tk.END, f"  {word}: {count} times\n")
            details_text.insert(tk.END, "\n")

        if semantic_analysis['related_queries']:
            details_text.insert(tk.END, "Related queries:\n", "subheader")
            for query, count in semantic_analysis['related_queries'][:10]:
                details_text.insert(tk.END, f"  {query} ({count} times)\n")
            details_text.insert(tk.END, "\n")

        if semantic_analysis['questions']:
            details_text.insert(tk.END, "People also ask:\n", "subheader")
            for question, count in semantic_analysis['questions'][:10]:
                details_text.insert(tk.END, f"  {question} ({count} times)\n")
            details_text.insert(tk.END, "\n")

    # Text formatting
    details_text.tag_configure("header", font=("Arial", 12, "bold"), foreground="blue")
    details_text.tag_configure("subheader", font=("Arial", 10, "bold"), foreground="darkblue")


def show_keyword_details(keyword):
    """
    Shows details of selected keyword
    """
    if not keyword or not current_serp_data or keyword not in current_serp_data:
        return

    # Clear details text field
    details_text.delete("1.0", tk.END)

    # Get keyword data
    freq = current_keywords_data.get(keyword, {}).get('frequency', 0) if current_keywords_data else 0
    prefix = current_keywords_data.get(keyword, {}).get('prefix', '') if current_keywords_data else ''

    # Get SERP results
    serp_results = current_serp_data.get(keyword, [])

    # Display information
    details_text.insert(tk.END, f"=== KEYWORD ===\n\n", "header")
    details_text.insert(tk.END, f"Query: {keyword}\n")
    details_text.insert(tk.END, f"Frequency: {freq:,}\n")
    if prefix:
        details_text.insert(tk.END, f"Prefix: {prefix}\n")
    details_text.insert(tk.END, "\n")

    # Display SERP results
    if serp_results:
        details_text.insert(tk.END, "Search results:\n", "subheader")
        for i, result in enumerate(serp_results, 1):
            details_text.insert(tk.END, f"  {i}. {result['title']}\n", "url_title")
            details_text.insert(tk.END, f"     URL: {result['url']}\n", "url")
            if result['description']:
                # Limit description length for readability
                description = result['description']
                if len(description) > 150:
                    description = description[:147] + "..."
                details_text.insert(tk.END, f"     {description}\n")
            details_text.insert(tk.END, "\n")

    # Display related queries
    if current_related_data and keyword in current_related_data:
        details_text.insert(tk.END, "Related queries:\n", "subheader")
        for i, query in enumerate(current_related_data[keyword][:10], 1):
            details_text.insert(tk.END, f"  {i}. {query}\n")
        details_text.insert(tk.END, "\n")

    # Display "People also ask"
    if current_paa_data and keyword in current_paa_data:
        details_text.insert(tk.END, "People also ask:\n", "subheader")
        for i, question in enumerate(current_paa_data[keyword][:10], 1):
            details_text.insert(tk.END, f"  {i}. {question}\n")
        details_text.insert(tk.END, "\n")

    # Text formatting
    details_text.tag_configure("header", font=("Arial", 12, "bold"), foreground="blue")
    details_text.tag_configure("subheader", font=("Arial", 10, "bold"), foreground="darkblue")
    details_text.tag_configure("url_title", font=("Arial", 9, "bold"))
    details_text.tag_configure("url", foreground="blue")


def search_in_clusters():
    """
    Searches for keyword in clusters
    """
    global search_results

    search_query = search_var.get().strip().lower()
    if not search_query or not current_clusters:
        return

    # Reset previous search results
    search_results = []

    # Find keywords containing search query
    for cluster_idx, cluster in enumerate(current_clusters):
        for keyword in cluster:
            if search_query in keyword.lower():
                search_results.append((cluster_idx, keyword))

    # Display search results
    if search_results:
        # Clear details text field
        details_text.delete("1.0", tk.END)

        details_text.insert(tk.END, f"=== SEARCH RESULTS FOR '{search_query}' ===\n\n", "header")
        details_text.insert(tk.END,
                            f"Found {len(search_results)} matches in {len(set(item[0] for item in search_results))} clusters\n\n")

        # Group results by cluster
        by_cluster = defaultdict(list)
        for cluster_idx, keyword in search_results:
            by_cluster[cluster_idx].append(keyword)

        # Output results
        for cluster_idx, keywords in by_cluster.items():
            details_text.insert(tk.END, f"Cluster {cluster_idx + 1} ({len(keywords)} matches):\n", "subheader")
            for kw in keywords:
                freq = current_keywords_data.get(kw, {}).get('frequency', 0) if current_keywords_data else 0
                details_text.insert(tk.END, f"  {kw} ({freq:,})\n")
            details_text.insert(tk.END, "\n")

        # Text formatting
        details_text.tag_configure("header", font=("Arial", 12, "bold"), foreground="blue")
        details_text.tag_configure("subheader", font=("Arial", 10, "bold"), foreground="darkblue")
    else:
        details_text.delete("1.0", tk.END)
        details_text.insert(tk.END, f"No results found for '{search_query}'.")


def reset_search():
    """
    Resets search
    """
    global search_results

    search_var.set("")
    search_results = []
    details_text.delete("1.0", tk.END)
    details_text.insert(tk.END, "Search results reset.")

def calculate_domain_stats(serp_data):
    """
    Calculates domain statistics to identify common domains in SERP results
    """
    domain_counter = Counter()

    for query, results in serp_data.items():
        for result in results:
            if result['url']:
                domain = urlparse(result['url']).netloc
                domain_counter[domain] += 1

    return domain_counter


def update_statistics():
    """
    Updates clustering statistics
    """
    if not current_clusters or not current_data:
        stats_text.delete("1.0", tk.END)
        stats_text.insert(tk.END, "No data to analyze. Please run clustering first.")
        return

    # Clear statistics text field
    stats_text.delete("1.0", tk.END)

    # Calculate cluster statistics
    eval_results = evaluate_clusters(current_clusters, current_data)

    # General statistics
    stats_text.insert(tk.END, "=== CLUSTERING STATISTICS ===\n\n", "header")
    stats_text.insert(tk.END, f"Number of clusters: {len(current_clusters)}\n")
    stats_text.insert(tk.END, f"Total keywords: {sum(len(c) for c in current_clusters)}\n")
    stats_text.insert(tk.END, f"Average intra-cluster similarity: {eval_results['avg_intra_sim']:.4f}\n")
    stats_text.insert(tk.END, f"Average inter-cluster similarity: {eval_results['avg_inter_sim']:.4f}\n")
    stats_text.insert(tk.END, f"Silhouette (clustering quality): {eval_results['silhouette']:.4f}\n\n")

    # Cluster size distribution
    stats_text.insert(tk.END, "=== CLUSTER SIZE DISTRIBUTION ===\n\n", "header")
    sizes = [len(cluster) for cluster in current_clusters]
    size_counts = Counter(sizes)

    for size, count in sorted(size_counts.items()):
        stats_text.insert(tk.END, f"Clusters of size {size}: {count}\n")

    stats_text.insert(tk.END, f"\nMinimum size: {min(sizes) if sizes else 0}\n")
    stats_text.insert(tk.END, f"Maximum size: {max(sizes) if sizes else 0}\n")
    stats_text.insert(tk.END, f"Average size: {sum(sizes) / len(sizes) if sizes else 0:.2f}\n\n")

    # Statistics by cluster
    stats_text.insert(tk.END, "=== INDIVIDUAL CLUSTER STATISTICS ===\n\n", "header")

    for i, stat in enumerate(eval_results['cluster_stats']):
        stats_text.insert(tk.END, f"Cluster {stat['id']}:\n", "subheader")
        stats_text.insert(tk.END, f"  Size: {stat['size']} keywords\n")
        stats_text.insert(tk.END, f"  Average frequency: {stat['avg_frequency']:.2f}\n")
        stats_text.insert(tk.END, f"  Total frequency: {stat['total_frequency']}\n")
        stats_text.insert(tk.END, f"  Intra-cluster similarity: {stat['intra_similarity']:.4f}\n")
        stats_text.insert(tk.END, f"  Central keyword: {stat['central_keyword']}\n\n")

    # Domain analysis
    if current_serp_data:
        stats_text.insert(tk.END, "=== DOMAIN STATISTICS ===\n\n", "header")
        domain_counter = calculate_domain_stats(current_serp_data)

        stats_text.insert(tk.END, "Top-10 most frequent domains:\n")
        for domain, count in domain_counter.most_common(10):
            stats_text.insert(tk.END, f"  {domain}: {count} occurrences\n")

    # Text formatting
    stats_text.tag_configure("header", font=("Arial", 12, "bold"), foreground="blue")
    stats_text.tag_configure("subheader", font=("Arial", 10, "bold"), foreground="darkblue")


def export_current_clusters(format_type):
    """
    Exports current clusters in selected format with progress indicator
    """
    if not current_clusters or not current_serp_data:
        messagebox.showwarning("Warning", "No data to export!")
        return

    # Создаем отдельное окно с прогрессом
    progress_window = tk.Toplevel(root)
    progress_window.title("Exporting Data")
    progress_window.geometry("400x150")
    progress_window.resizable(False, False)
    progress_window.transient(root)
    progress_window.grab_set()  # Делаем модальным

    # Центрируем окно
    window_width = 400
    window_height = 150
    screen_width = progress_window.winfo_screenwidth()
    screen_height = progress_window.winfo_screenheight()
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    progress_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # Добавляем элементы интерфейса
    ttk.Label(progress_window, text=f"Exporting data to {format_type.upper()}...",
              font=("Segoe UI", 12)).pack(pady=(15, 10))

    export_progress = ttk.Progressbar(progress_window, orient=tk.HORIZONTAL,
                                      length=350, mode='indeterminate')
    export_progress.pack(pady=10, padx=25)

    status_label = ttk.Label(progress_window, text="Preparing data...", font=("Segoe UI", 9))
    status_label.pack(pady=5)

    cancel_button = ttk.Button(progress_window, text="Cancel", command=progress_window.destroy)
    cancel_button.pack(pady=10)

    # Запускаем индикатор прогресса
    export_progress.start(15)

    def update_status(message):
        status_label.config(text=message)
        progress_window.update_idletasks()

    # Экспорт данных в отдельном потоке
    def export_thread():
        try:
            if format_type == "excel":
                # Выбор имени файла
                filename = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    title="Export to Excel"
                )

                if filename:
                    update_status("Creating Excel file...")
                    include_details = export_detail.get()
                    result = export_clusters_to_excel(current_clusters, current_serp_data,
                                                      filename, include_details)

                    # Закрываем окно прогресса и показываем сообщение об успехе
                    progress_window.after(500, lambda: progress_window.destroy())
                    messagebox.showinfo("Export", f"Data successfully exported to {result}")
                else:
                    progress_window.destroy()

            elif format_type == "csv":
                # Выбор имени файла
                filename = filedialog.asksaveasfilename(
                    defaultextension=".zip",
                    filetypes=[("ZIP files", "*.zip")],
                    title="Export to CSV (ZIP archive)"
                )

                if filename:
                    update_status("Creating CSV files and packaging to ZIP...")
                    result = export_clusters_to_csv(current_clusters, current_serp_data, filename)

                    # Закрываем окно прогресса и показываем сообщение об успехе
                    progress_window.after(500, lambda: progress_window.destroy())
                    messagebox.showinfo("Export", f"Data successfully exported to {result}")
                else:
                    progress_window.destroy()

            elif format_type == "json":
                # Выбор имени файла
                filename = filedialog.asksaveasfilename(
                    defaultextension=".json",
                    filetypes=[("JSON files", "*.json")],
                    title="Export to JSON"
                )

                if filename:
                    update_status("Creating JSON file...")
                    result = export_clusters_to_json(current_clusters, current_serp_data, filename)

                    # Закрываем окно прогресса и показываем сообщение об успехе
                    progress_window.after(500, lambda: progress_window.destroy())
                    messagebox.showinfo("Export", f"Data successfully exported to {result}")
                else:
                    progress_window.destroy()

        except Exception as e:
            # При ошибке закрываем окно прогресса и показываем сообщение об ошибке
            progress_window.destroy()
            messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")

    # Запускаем экспорт в отдельном потоке
    export_thread = threading.Thread(target=export_thread)
    export_thread.daemon = True
    export_thread.start()


def show_about():
    """
    Shows program information
    """
    about_text = """
SEO Keyword Clustering v1.0

This program allows you to cluster keywords based on 
SERP (Search Engine Results Page) overlap for Google.

It helps with:
- Content planning
- Website structure optimization
- Identifying semantic relationships
- SEO strategy development

This tool is similar to functionality in:
- SERPstat
- SEMrush
- SE Ranking
- Rush Analytics
- Keyword Clusterizer

© 2024 SEO Tools
"""
    messagebox.showinfo("About", about_text)


def show_help():
    """
    Shows help information
    """
    help_text = """
Keyword Clustering Tool User Guide:

1. Getting Started:
   - Load data from database using "File" -> "Load Data"
   - Apply filters in the "Filtering" tab if needed

2. Clustering Setup:
   - Choose clustering type: soft or hard
   - Set URL overlap threshold (lower = bigger clusters)
   - Set maximum SERP position to analyze

3. Running Clustering:
   - Click "Run Clustering" button
   - Wait for the process to complete
   - Results will appear in the clusters tree

4. Analyzing Results:
   - View individual clusters by selecting them in the tree
   - Explore cluster details or keyword details
   - Use search to find specific keywords

5. Exporting Results:
   - Export to Excel, CSV, JSON or HTML via the "File" menu
   - Configure export details in clustering tab

For best results:
- Use soft clustering with threshold 2-3 for content planning
- Use hard clustering with threshold 3-5 for strict topic separation
- Analyze common URLs to identify potential content gaps
- Review "People also ask" for content ideas

For additional help, contact support.
"""
    help_window = tk.Toplevel(root)
    help_window.title("Help")
    help_window.geometry("600x500")

    help_text_widget = tk.Text(help_window, wrap=tk.WORD, padx=10, pady=10)
    help_text_widget.pack(fill=tk.BOTH, expand=True)

    help_text_widget.insert(tk.END, help_text)
    help_text_widget.config(state=tk.DISABLED)


def run_processing_in_thread(func, callback=None, timeout=300):
    """
    Runs function in separate thread with optional timeout
    """
    global cancellable_thread

    # Event for timeout signal
    timeout_event = threading.Event()

    def timeout_handler():
        """Function executed when timeout occurs"""
        if not timeout_event.is_set():
            timeout_event.set()
            root.after(100, lambda: messagebox.showinfo(
                "Timeout", f"Operation canceled due to timeout ({timeout} seconds)"))
            stop_processing()

    def wrapper():
        try:
            # Set timer for timeout
            timer = threading.Timer(timeout, timeout_handler)
            timer.daemon = True
            timer.start()

            result = func()

            # Cancel timer if operation completed
            timer.cancel()
            timeout_event.set()

            # Execute callback in main thread if operation completed
            if callback and not stop_clustering:
                root.after(100, lambda: callback(result))
        except Exception as e:
            # Ensure timer is canceled
            if 'timer' in locals():
                timer.cancel()
            timeout_event.set()

            messagebox.showerror("Processing Error", f"An error occurred: {str(e)}")

    cancellable_thread = threading.Thread(target=wrapper)
    cancellable_thread.daemon = True
    cancellable_thread.start()


def enable_copying_for_all_text_widgets():
    """
    Разрешает копирование текста из всех текстовых виджетов
    """

    def bind_copy_paste(widget):
        widget.bind("<Control-c>", lambda event: widget.event_generate("<<Copy>>"))
        widget.bind("<Control-a>", lambda event: widget.event_generate("<<SelectAll>>"))

    for widget in [details_text, stats_text]:
        if widget:
            bind_copy_paste(widget)
            # Разрешаем выделение и копирование текста правым кликом
            widget.config(state=tk.NORMAL)

    # Применяем к дополнительным текстовым виджетам
    if excluded_domains_text:
        bind_copy_paste(excluded_domains_text)


def create_context_menu():
    """
    Создает контекстное меню для TreeView с кластерами и ключевыми запросами
    """
    global clusters_tree

    # Создаем контекстное меню
    context_menu = tk.Menu(clusters_tree, tearoff=0)

    # Добавляем пункты меню
    context_menu.add_command(label="TOP URLs", command=lambda: show_top_urls())
    context_menu.add_command(label="TOP Titles", command=lambda: show_top_titles())
    context_menu.add_command(label="Related Queries", command=lambda: show_related_queries())
    context_menu.add_command(label="People Also Ask", command=lambda: show_people_also_ask())

    # Привязываем контекстное меню к правому клику на TreeView
    def show_context_menu(event):
        # Получаем выбранный элемент
        item = clusters_tree.identify_row(event.y)
        if item:
            # Выбираем элемент, на котором был сделан клик
            clusters_tree.selection_set(item)
            # Показываем контекстное меню
            context_menu.post(event.x_root, event.y_root)

    # Привязываем правый клик к показу контекстного меню
    clusters_tree.bind("<Button-3>", show_context_menu)

    return context_menu


def show_top_urls():
    """
    Показывает TOP URLs для выбранного кластера или ключевого запроса
    """
    # Получаем выбранный элемент
    selected_items = clusters_tree.selection()
    if not selected_items:
        return

    selected_item = selected_items[0]

    # Определяем, выбран ли кластер или ключевой запрос
    parent_id = clusters_tree.parent(selected_item)

    if not parent_id:  # Это кластер
        cluster_id = int(clusters_tree.item(selected_item)['values'][0])
        show_cluster_top_urls(cluster_id - 1)  # -1 потому что индексация начинается с 0
    else:  # Это ключевой запрос
        keyword = clusters_tree.item(selected_item)['text']
        show_keyword_top_urls(keyword)


def show_top_titles():
    """
    Показывает TOP Titles для выбранного кластера или ключевого запроса
    """
    # Получаем выбранный элемент
    selected_items = clusters_tree.selection()
    if not selected_items:
        return

    selected_item = selected_items[0]

    # Определяем, выбран ли кластер или ключевой запрос
    parent_id = clusters_tree.parent(selected_item)

    if not parent_id:  # Это кластер
        cluster_id = int(clusters_tree.item(selected_item)['values'][0])
        show_cluster_top_titles(cluster_id - 1)
    else:  # Это ключевой запрос
        keyword = clusters_tree.item(selected_item)['text']
        show_keyword_top_titles(keyword)


def show_related_queries():
    """
    Показывает связанные запросы для выбранного кластера или ключевого запроса
    """
    # Получаем выбранный элемент
    selected_items = clusters_tree.selection()
    if not selected_items:
        return

    selected_item = selected_items[0]

    # Определяем, выбран ли кластер или ключевой запрос
    parent_id = clusters_tree.parent(selected_item)

    if not parent_id:  # Это кластер
        cluster_id = int(clusters_tree.item(selected_item)['values'][0])
        show_cluster_related_queries(cluster_id - 1)
    else:  # Это ключевой запрос
        keyword = clusters_tree.item(selected_item)['text']
        show_keyword_related_queries(keyword)


def show_people_also_ask():
    """
    Показывает "People Also Ask" вопросы для выбранного кластера или ключевого запроса
    """
    # Получаем выбранный элемент
    selected_items = clusters_tree.selection()
    if not selected_items:
        return

    selected_item = selected_items[0]

    # Определяем, выбран ли кластер или ключевой запрос
    parent_id = clusters_tree.parent(selected_item)

    if not parent_id:  # Это кластер
        cluster_id = int(clusters_tree.item(selected_item)['values'][0])
        show_cluster_people_also_ask(cluster_id - 1)
    else:  # Это ключевой запрос
        keyword = clusters_tree.item(selected_item)['text']
        show_keyword_people_also_ask(keyword)


def show_cluster_top_urls(cluster_index):
    """
    Показывает TOP URLs для выбранного кластера
    """
    if not current_clusters or cluster_index >= len(current_clusters):
        return

    cluster = current_clusters[cluster_index]

    # Создаем список всех URL из ТОП-10 для запросов в кластере
    urls_data = []
    for kw in cluster:
        if kw in current_serp_data:
            for result in current_serp_data[kw]:
                if result['position'] <= 10:  # Только ТОП-10
                    urls_data.append({
                        'keyword': kw,
                        'position': result['position'],
                        'url': result['url'],
                        'normalized_url': result['normalized_url']
                    })

    # Если данных нет, показываем сообщение
    if not urls_data:
        messagebox.showinfo("Top URLs", "No TOP-10 URLs found for this cluster")
        return

    # Показываем данные в новом окне
    show_data_window("TOP-10 URLs for Cluster #" + str(cluster_index + 1),
                     urls_data,
                     ['keyword', 'position', 'url'],
                     ['Keyword', 'Position', 'URL'])


def show_cluster_top_titles(cluster_index):
    """
    Показывает TOP Titles для выбранного кластера
    """
    if not current_clusters or cluster_index >= len(current_clusters):
        return

    cluster = current_clusters[cluster_index]

    # Создаем список всех заголовков из ТОП-10 для запросов в кластере
    titles_data = []
    for kw in cluster:
        if kw in current_serp_data:
            for result in current_serp_data[kw]:
                if result['position'] <= 10:  # Только ТОП-10
                    titles_data.append({
                        'keyword': kw,
                        'position': result['position'],
                        'title': result['title']
                    })

    # Если данных нет, показываем сообщение
    if not titles_data:
        messagebox.showinfo("Top Titles", "No TOP-10 titles found for this cluster")
        return

    # Показываем данные в новом окне
    show_data_window("TOP-10 Titles for Cluster #" + str(cluster_index + 1),
                     titles_data,
                     ['keyword', 'position', 'title'],
                     ['Keyword', 'Position', 'Title'])


def show_cluster_related_queries(cluster_index):
    """
    Показывает связанные запросы для выбранного кластера
    """
    if not current_clusters or cluster_index >= len(current_clusters) or not current_related_data:
        return

    cluster = current_clusters[cluster_index]

    # Создаем список всех связанных запросов для кластера
    related_data = []
    for kw in cluster:
        if kw in current_related_data:
            for related in current_related_data[kw]:
                related_data.append({
                    'keyword': kw,
                    'related_query': related
                })

    # Если данных нет, показываем сообщение
    if not related_data:
        messagebox.showinfo("Related Queries", "No related queries found for this cluster")
        return

    # Показываем данные в новом окне
    show_data_window("Related Queries for Cluster #" + str(cluster_index + 1),
                     related_data,
                     ['keyword', 'related_query'],
                     ['Keyword', 'Related Query'])


def show_cluster_people_also_ask(cluster_index):
    """
    Показывает "People Also Ask" вопросы для выбранного кластера
    """
    if not current_clusters or cluster_index >= len(current_clusters) or not current_paa_data:
        return

    cluster = current_clusters[cluster_index]

    # Создаем список всех "People Also Ask" вопросов для кластера
    paa_data = []
    for kw in cluster:
        if kw in current_paa_data:
            for question in current_paa_data[kw]:
                paa_data.append({
                    'keyword': kw,
                    'question': question
                })

    # Если данных нет, показываем сообщение
    if not paa_data:
        messagebox.showinfo("People Also Ask", "No 'People Also Ask' questions found for this cluster")
        return

    # Показываем данные в новом окне
    show_data_window("People Also Ask Questions for Cluster #" + str(cluster_index + 1),
                     paa_data,
                     ['keyword', 'question'],
                     ['Keyword', 'Question'])


def show_keyword_top_urls(keyword):
    """
    Показывает TOP URLs для выбранного ключевого запроса
    """
    if not keyword or not current_serp_data or keyword not in current_serp_data:
        return

    # Создаем список URL из ТОП-10 для запроса
    urls_data = []
    for result in current_serp_data[keyword]:
        if result['position'] <= 10:  # Только ТОП-10
            urls_data.append({
                'position': result['position'],
                'url': result['url']
            })

    # Если данных нет, показываем сообщение
    if not urls_data:
        messagebox.showinfo("Top URLs", f"No TOP-10 URLs found for keyword '{keyword}'")
        return

    # Показываем данные в новом окне
    show_data_window(f"TOP-10 URLs for Keyword: {keyword}",
                     urls_data,
                     ['position', 'url'],
                     ['Position', 'URL'])


def show_keyword_top_titles(keyword):
    """
    Показывает TOP Titles для выбранного ключевого запроса
    """
    if not keyword or not current_serp_data or keyword not in current_serp_data:
        return

    # Создаем список заголовков из ТОП-10 для запроса
    titles_data = []
    for result in current_serp_data[keyword]:
        if result['position'] <= 10:  # Только ТОП-10
            titles_data.append({
                'position': result['position'],
                'title': result['title']
            })

    # Если данных нет, показываем сообщение
    if not titles_data:
        messagebox.showinfo("Top Titles", f"No TOP-10 titles found for keyword '{keyword}'")
        return

    # Показываем данные в новом окне
    show_data_window(f"TOP-10 Titles for Keyword: {keyword}",
                     titles_data,
                     ['position', 'title'],
                     ['Position', 'Title'])


def show_keyword_related_queries(keyword):
    """
    Показывает связанные запросы для выбранного ключевого запроса
    """
    if not keyword or not current_related_data or keyword not in current_related_data:
        messagebox.showinfo("Related Queries", f"No related queries found for keyword '{keyword}'")
        return

    # Создаем список связанных запросов
    related_data = []
    for related in current_related_data[keyword]:
        related_data.append({
            'related_query': related
        })

    # Показываем данные в новом окне
    show_data_window(f"Related Queries for Keyword: {keyword}",
                     related_data,
                     ['related_query'],
                     ['Related Query'])


def show_keyword_people_also_ask(keyword):
    """
    Показывает "People Also Ask" вопросы для выбранного ключевого запроса
    """
    if not keyword or not current_paa_data or keyword not in current_paa_data:
        messagebox.showinfo("People Also Ask", f"No 'People Also Ask' questions found for keyword '{keyword}'")
        return

    # Создаем список "People Also Ask" вопросов
    paa_data = []
    for question in current_paa_data[keyword]:
        paa_data.append({
            'question': question
        })

    # Показываем данные в новом окне
    show_data_window(f"People Also Ask Questions for Keyword: {keyword}",
                     paa_data,
                     ['question'],
                     ['Question'])


# 6. Общая функция для показа данных в отдельном окне

def show_data_window(title, data, fields, headers):
    """
    Показывает данные в отдельном окне с возможностью копирования

    Args:
        title (str): Заголовок окна
        data (list): Список словарей с данными
        fields (list): Список полей для отображения из словарей
        headers (list): Заголовки для отображаемых полей
    """
    # Создаем новое окно
    data_window = tk.Toplevel(root)
    data_window.title(title)
    data_window.geometry("800x600")
    data_window.minsize(500, 300)

    # Добавляем текстовое поле для отображения данных
    data_text = tk.Text(data_window, wrap=tk.WORD, padx=10, pady=10)
    data_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # Добавляем скроллбар
    scrollbar = ttk.Scrollbar(data_text, command=data_text.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    data_text.configure(yscrollcommand=scrollbar.set)

    # Заполняем данными
    # Добавляем заголовки
    header_line = "\t".join(headers) + "\n"
    data_text.insert(tk.END, header_line, "header")
    data_text.insert(tk.END, "-" * len(header_line) + "\n\n")

    # Добавляем строки данных
    for item in data:
        row_data = []
        for field in fields:
            row_data.append(str(item.get(field, "")))
        data_text.insert(tk.END, "\t".join(row_data) + "\n")

    # Форматирование
    data_text.tag_configure("header", font=("Arial", 10, "bold"))

    # Разрешаем копирование текста
    def bind_copy_paste(widget):
        widget.bind("<Control-c>", lambda event: widget.event_generate("<<Copy>>"))
        widget.bind("<Control-a>", lambda event: widget.event_generate("<<SelectAll>>"))

    bind_copy_paste(data_text)

    # Создаем кнопку копирования всего содержимого
    def copy_all_data():
        data_text.tag_add(tk.SEL, "1.0", tk.END)
        data_text.event_generate("<<Copy>>")
        data_text.tag_remove(tk.SEL, "1.0", tk.END)
        messagebox.showinfo("Copy", "All data copied to clipboard")

    copy_button = ttk.Button(data_window, text="Copy All", command=copy_all_data)
    copy_button.pack(pady=10)


def initialize_gui_updated():
    """
    Обновленная функция инициализации GUI с расширенными возможностями копирования
    """
    # Вызываем исходную функцию инициализации
    initialize_gui()

    # Добавляем расширенные возможности копирования для всех текстовых виджетов
    enable_copying_for_all_text_widgets()

    # Улучшаем панель деталей кластера
    enhance_cluster_details_panel()

    # Создаем контекстное меню для дерева кластеров
    create_context_menu()

    # Переопределяем функцию обработки выбора кластера для вызова улучшенной версии
    clusters_tree.bind("<<TreeviewSelect>>", lambda e: on_cluster_select_enhanced(e))


def on_cluster_select_enhanced(event):
    """
    Улучшенная функция обработки выбора элемента в дереве кластеров,
    которая использует расширенные функции отображения
    """
    selected_items = clusters_tree.selection()
    if not selected_items:
        return

    selected_item = selected_items[0]

    # Проверяем, это кластер или ключевое слово
    parent_id = clusters_tree.parent(selected_item)

    if not parent_id:  # Это кластер
        cluster_id = int(clusters_tree.item(selected_item)['values'][0])
        show_cluster_details_enhanced(cluster_id - 1)  # -1 потому что индексация начинается с 0
    else:  # Это ключевое слово
        keyword = clusters_tree.item(selected_item)['text']
        show_keyword_details(keyword)  # Можно оставить исходную функцию или создать аналогичную улучшенную


def enhance_cluster_details_panel():
    """
    Enhances the Cluster Details panel interface, adding copy functionality
    and integrating copy buttons into the GUI
    """
    global details_text, details_frame

    # 1. Create buttons for copying in the Cluster Details panel
    buttons_frame = ttk.Frame(details_frame)
    buttons_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=5)

    # Button for copying all content
    copy_all_button = ttk.Button(
        buttons_frame,
        text="Copy All",
        command=lambda: copy_details_content("all")
    )
    copy_all_button.pack(side=tk.LEFT, padx=5)

    # Button for copying only keywords
    copy_keywords_button = ttk.Button(
        buttons_frame,
        text="Copy Keywords",
        command=lambda: copy_details_content("keywords")
    )
    copy_keywords_button.pack(side=tk.LEFT, padx=5)

    # Button for copying common URLs
    copy_urls_button = ttk.Button(
        buttons_frame,
        text="Copy URLs",
        command=lambda: copy_details_content("urls")
    )
    copy_urls_button.pack(side=tk.LEFT, padx=5)

    # 2. Allow selecting and copying text directly in the text field
    details_text.configure(state=tk.NORMAL)  # Ensure text can be selected

    # Bind standard keyboard shortcuts
    details_text.bind("<Control-a>", lambda e: select_all_text(details_text))
    details_text.bind("<Control-c>", lambda e: copy_selected_text(details_text))

    # Add right-click context menu
    create_details_context_menu()


def create_details_context_menu():
    """
    Creates a context menu for the cluster details text field
    """
    context_menu = tk.Menu(details_text, tearoff=0)
    context_menu.add_command(label="Copy", command=lambda: copy_selected_text(details_text))
    context_menu.add_command(label="Select All", command=lambda: select_all_text(details_text))
    context_menu.add_separator()
    context_menu.add_command(label="Copy All Content", command=lambda: copy_details_content("all"))
    context_menu.add_command(label="Copy Keywords Only", command=lambda: copy_details_content("keywords"))
    context_menu.add_command(label="Copy URLs Only", command=lambda: copy_details_content("urls"))

    # Bind context menu to right-click
    details_text.bind("<Button-3>", lambda e: show_context_menu(e, context_menu))


def show_context_menu(event, menu):
    """
    Показывает контекстное меню в указанной позиции
    """
    menu.post(event.x_root, event.y_root)


def select_all_text(text_widget):
    """
    Выделяет весь текст в текстовом виджете
    """
    text_widget.tag_add(tk.SEL, "1.0", tk.END)
    text_widget.mark_set(tk.INSERT, "1.0")
    text_widget.see(tk.INSERT)
    return "break"  # Предотвращает стандартное поведение виджета


def copy_selected_text(text_widget):
    """
    Копирует выбранный текст в буфер обмена
    """
    try:
        selected_text = text_widget.get(tk.SEL_FIRST, tk.SEL_LAST)
        root.clipboard_clear()
        root.clipboard_append(selected_text)
    except tk.TclError:
        # Если текст не выбран, ничего не делаем
        pass
    return "break"  # Предотвращает стандартное поведение виджета


def copy_details_content(content_type="all"):
    """
    Copies the content of Cluster Details to clipboard
    based on the selected content type

    Args:
        content_type (str): Type of content to copy
            - "all": all content
            - "keywords": only keywords
            - "urls": only URLs
    """
    all_content = details_text.get("1.0", tk.END)

    if content_type == "all":
        # Copy all content
        text_to_copy = all_content
    elif content_type == "keywords":
        # Extract only lines with keywords
        lines = all_content.split('\n')
        keywords_lines = []

        # Find the beginning of keywords section
        keywords_section_started = False
        for line in lines:
            if "Keywords in cluster:" in line:
                keywords_section_started = True
                continue

            if keywords_section_started:
                # If we hit an empty line, the section has ended
                if line.strip() == "":
                    break

                # Add the line to our list, removing leading spaces and frequency info
                kw_line = line.strip()
                if kw_line and kw_line.startswith("  "):
                    # Remove frequency if it's in parentheses
                    if "(" in kw_line:
                        kw_line = kw_line.split("(")[0].strip()
                    keywords_lines.append(kw_line)

        text_to_copy = "\n".join(keywords_lines)
    elif content_type == "urls":
        # Extract only lines with URLs
        lines = all_content.split('\n')
        urls_lines = []

        # Find beginning of URLs section
        urls_section_started = False
        for line in lines:
            if "Common URLs" in line:
                urls_section_started = True
                continue

            if urls_section_started:
                # If we hit an empty line, the section has ended
                if line.strip() == "":
                    break

                # Add URL, removing leading spaces and number in list
                url_line = line.strip()
                if url_line and url_line.startswith("  "):
                    # Remove URL number in list (if specified)
                    if ". " in url_line:
                        url_line = url_line.split(". ", 1)[1].strip()
                    urls_lines.append(url_line)

        text_to_copy = "\n".join(urls_lines)

    # Copy text to clipboard
    if text_to_copy and text_to_copy.strip():
        root.clipboard_clear()
        root.clipboard_append(text_to_copy)
        messagebox.showinfo("Copy", f"Content copied to clipboard ({content_type})")
    else:
        messagebox.showinfo("Copy", f"Nothing to copy. Make sure to select a cluster first.")


def show_cluster_details_enhanced(cluster_index):
    """
    Enhanced version of cluster details display function with better formatting
    to facilitate copying
    """
    if not current_clusters or cluster_index >= len(current_clusters):
        return

    cluster = current_clusters[cluster_index]

    # Clear text field
    details_text.delete("1.0", tk.END)

    # Determine central keyword
    central_kw = get_central_keyword(cluster, current_keywords_data)

    # Collect common URLs information
    common_urls = None
    if current_serp_data:
        url_counts = Counter()
        for kw in cluster:
            if kw in current_serp_data:
                for result in current_serp_data[kw]:
                    if result['normalized_url']:
                        url_counts[result['normalized_url']] += 1

        # Find URLs present in at least half of the queries
        threshold = max(2, len(cluster) / 2)
        common_urls = [url for url, count in url_counts.most_common(10) if count >= threshold]

    # Analyze semantic similarity
    semantic_analysis = None
    if current_related_data or current_paa_data:
        try:
            semantic_analysis = analyze_keyword_semantics([cluster],
                                                       current_related_data,
                                                       current_paa_data)[1]
        except:
            semantic_analysis = {
                'common_words': {},
                'related_queries': [],
                'questions': []
            }

    # Collect cluster statistics
    total_freq = 0
    if current_keywords_data:
        for kw in cluster:
            if kw in current_keywords_data:
                total_freq += current_keywords_data[kw].get('frequency', 0)

    # Display information with updated formatting
    details_text.insert(tk.END, f"=== CLUSTER {cluster_index + 1} ===\n\n", "header")
    details_text.insert(tk.END, f"Size: {len(cluster)} keywords\n")
    details_text.insert(tk.END, f"Total frequency: {total_freq:,}\n")

    if central_kw:
        central_freq = current_keywords_data.get(central_kw, {}).get('frequency', 0) if current_keywords_data else 0
        details_text.insert(tk.END, f"Central keyword: {central_kw} (frequency: {central_freq:,})\n\n")

    # Display keywords - this is an important section for copying
    details_text.insert(tk.END, "Keywords in cluster:\n", "subheader")
    sorted_keywords = sorted(cluster,
                             key=lambda kw: current_keywords_data.get(kw, {}).get('frequency', 0)
                             if current_keywords_data else 0,
                             reverse=True)

    for kw in sorted_keywords:  # Removed limit for easier copying of all keywords
        freq = current_keywords_data.get(kw, {}).get('frequency', 0) if current_keywords_data else 0
        details_text.insert(tk.END, f"  {kw} ({freq:,})\n")

    details_text.insert(tk.END, "\n")

    # Display common URLs - another important section for copying
    if common_urls:
        details_text.insert(tk.END, "Common URLs (present in at least half the keywords):\n", "subheader")
        for i, url in enumerate(common_urls, 1):
            details_text.insert(tk.END, f"  {i}. {url}\n")
        details_text.insert(tk.END, "\n")
    else:
        details_text.insert(tk.END, "Common URLs (present in at least half the keywords):\n", "subheader")
        details_text.insert(tk.END, "  No common URLs found\n\n")

    # Display semantic analysis
    if semantic_analysis:
        if semantic_analysis.get('common_words'):
            details_text.insert(tk.END, "Most frequent words in queries:\n", "subheader")
            for word, count in list(semantic_analysis['common_words'].items())[:10]:
                details_text.insert(tk.END, f"  {word}: {count} times\n")
            details_text.insert(tk.END, "\n")

        if semantic_analysis.get('related_queries'):
            details_text.insert(tk.END, "Related queries:\n", "subheader")
            for query, count in semantic_analysis['related_queries'][:10]:
                details_text.insert(tk.END, f"  {query} ({count} times)\n")
            details_text.insert(tk.END, "\n")

        if semantic_analysis.get('questions'):
            details_text.insert(tk.END, "People also ask:\n", "subheader")
            for question, count in semantic_analysis['questions'][:10]:
                details_text.insert(tk.END, f"  {question} ({count} times)\n")
            details_text.insert(tk.END, "\n")

    # Text formatting
    details_text.tag_configure("header", font=("Arial", 12, "bold"), foreground="blue")
    details_text.tag_configure("subheader", font=("Arial", 10, "bold"), foreground="darkblue")

    # Make sure text can be selected and copied
    details_text.configure(state=tk.NORMAL)

# ===================== Main Function =====================

def main():
    """
    Main program function
    """
    # Initialize UI with updated features
    initialize_gui_updated()

    # Run main loop
    root.mainloop()


# Start program
if __name__ == "__main__":
    main()